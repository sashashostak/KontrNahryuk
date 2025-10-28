#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
process_dodatok10.py - Обробка Додатку 10 (адаптація VBA логіки)

Функціонал:
1. Рекурсивний пошук всіх файлів з назвами "Дод.10", "Дод 10", "Додаток 10"
2. Збір даних з колонок B:FP (2..172) з всіх знайдених файлів
3. Нормалізація назв підрозділів (S-2, S-4, S-6, РСпП, тощо)
4. Глобальна дедуплікація (між усіма файлами)
5. Очищення цільового файлу (нижче блоку "УПР")
6. Запис у фіксованому порядку підрозділів
7. Збереження формул у колонках H(8), I(9), ET(150)

Підтримувані шаблони назв файлів:
- Дод.10_ВЗ 18.10.25.xlsx
- Дод 10_РВП 11.10.2025.xlsx
- Додаток 10_МБ 18.10.25.xlsx
- і т.д. (дата може змінюватися)

Константи:
- TOKEN_PATTERNS - regex шаблони для пошуку файлів
- COL_H, COL_I, COL_ET = 8, 9, 150 - колонки з формулами (НЕ перезаписувати)
- MAX_ROW = 900 - межа запису
"""

import os
import sys
import re
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Set
from collections import OrderedDict
from functools import partial

# Автоматичний flush для всіх print()
print = partial(print, flush=True)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
except ImportError:
    print("ERROR: openpyxl не встановлено. Встановіть: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Імпорт модуля оновлення статусу FO
try:
    from update_fo_status import update_fo_status
except ImportError:
    print("⚠️ WARNING: update_fo_status.py не знайдено. Оновлення статусу FO буде пропущено.", file=sys.stderr)
    update_fo_status = None


# ============== КОНСТАНТИ ==============
# Шаблони пошуку: файли з "Дод.10", "Дод 10", "Додаток 10" у назві
TOKEN_PATTERNS = [
    r"дод\.?\s*10",      # Дод.10 або Дод 10
    r"додаток\s*10",     # Додаток 10
    r"дод\.?\s*10_",     # Дод.10_ або Дод 10_
]

COL_H = 8
COL_I = 9
COL_ET = 150
MAX_ROW = 900
FORMULA_COLS = {COL_H, COL_I, COL_ET}  # Зберігаємо формули у H, I, ET

# Діапазон колонок B:FP = 2..172
COL_START = 2  # B
COL_END = 172  # FP

# Порядок підрозділів (після УПР)
UNITS_ORDER = [
    "S-2", "S-4", "S-6",
    "1РСпП", "2РСпП", "3РСпП",
    "РВП", "мінометна батарея", "РБпС", "ВРЕБ", "ВРСП",
    "відділення інструкторів", "Взвод зв'язку", "РМТЗ", "Медичний пункт"
]


# ============== ДОПОМІЖНІ ФУНКЦІЇ ДЛЯ ЗЛИТИХ КОМІРОК ==============

def get_writable_cell(ws: "openpyxl.worksheet.worksheet.Worksheet", row: int, col: int):
    """Повертає комірку, гарантувавши що вона не є злитою."""
    cell = ws.cell(row, col)

    if isinstance(cell, MergedCell):
        target_range = None
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                target_range = merged_range
                break

        if target_range is not None:
            ws.unmerge_cells(str(target_range))
            cell = ws.cell(row, col)

    return cell


def set_cell_value(ws: "openpyxl.worksheet.worksheet.Worksheet", row: int, col: int, value):
    """Встановлює значення в комірку, попередньо роз'єднавши злиття якщо потрібно."""
    cell = get_writable_cell(ws, row, col)
    cell.value = value


# ============== НОРМАЛІЗАЦІЯ НАЗВ ПІДРОЗДІЛІВ ==============

def normalize_spaces(s: str) -> str:
    """Нормалізує всі види пробілів"""
    s = s.replace('\xa0', ' ')  # неrozривний пробіл
    s = s.replace('\t', ' ')
    s = re.sub(r'\s+', ' ', s)  # багаторазові пробіли -> один
    return s.strip()


def canon_unit(raw: str) -> Optional[str]:
    """
    Повертає канонічну назву підрозділу або None якщо не розпізнано.
    Ігнорує регістр, нормалізує пробіли.
    """
    s = normalize_spaces(raw).upper()
    if not s:
        return None
    
    # S-2 (Штаб 2, штаБ-2, тощо)
    if 'ШТАБ' in s and ('2' in s or 'S-2' in s):
        return "S-2"
    
    # S-4
    if 'S-4' in s:
        return "S-4"
    
    # S-6 (зв'язок/ІТ)
    if 'S-6' in s:
        return "S-6"
    
    # РСпП (1/2/3) - з пробілом або без
    if 'РСПП' in s:
        # Видаляємо всі пробіли для перевірки
        s_no_space = s.replace(' ', '')
        if s_no_space.startswith('1') or s.startswith('1'):
            return "1РСпП"
        elif s_no_space.startswith('2') or s.startswith('2'):
            return "2РСпП"
        elif s_no_space.startswith('3') or s.startswith('3'):
            return "3РСпП"
    
    # РВП
    if s == 'РВП':
        return "РВП"
    
    # Мінометна батарея
    if 'МІНОМЕТН' in s:
        return "мінометна батарея"
    
    # РБпС (можливі варіанти написання)
    if s in ('РБПС', 'РБПC'):  # англійська C vs кирилична С
        return "РБпС"
    
    # ВРЕБ
    if s == 'ВРЕБ':
        return "ВРЕБ"
    
    # ВРСП
    if s == 'ВРСП':
        return "ВРСП"
    
    # Відділення інструкторів
    if 'ІНСТРУКТОР' in s:
        return "відділення інструкторів"
    
    # Взвод зв'язку
    if 'ВЗВОД' in s and any(x in s for x in ("ЗВ'ЯЗ", 'ЗВЯЗ', "ЗВ'ЯЗ")):
        return "Взвод зв'язку"
    
    # РМТЗ
    if s == 'РМТЗ':
        return "РМТЗ"
    
    # Медичний пункт
    if 'МЕД' in s:
        return "Медичний пункт"
    
    return None


# ============== ПОШУК АРКУША ТА БЛОКУ "УПР" ==============

def starts_with_upr(value) -> bool:
    """Перевіряє чи значення починається з "УПР" (ігноруючи пробіли і регістр)"""
    if value is None:
        return False
    s = normalize_spaces(str(value)).upper().replace(' ', '')
    return s.startswith('УПР')


def detect_target_sheet(wb: openpyxl.Workbook) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
    """Знаходить аркуш з блоком "УПР" у колонці B (рядки 10..MAX_ROW)"""
    for ws in wb.worksheets:
        # Пропускаємо LOG аркуш при пошуку
        if ws.title == 'LOG':
            continue
        for row in range(10, min(ws.max_row + 1, MAX_ROW + 1)):
            cell_value = ws.cell(row, 2).value  # колонка B
            if starts_with_upr(cell_value):
                return ws
    return None


def find_upr_end(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """
    Знаходить останній рядок блоку "УПР".
    Дозволяє 1 порожній рядок всередині "гребінки" УПР.
    """
    start_row = 0
    for row in range(10, min(ws.max_row + 1, MAX_ROW + 1)):
        if starts_with_upr(ws.cell(row, 2).value):
            start_row = row
            break
    
    if start_row == 0:
        return 0
    
    upr_end = start_row
    while upr_end < MAX_ROW:
        next_val = ws.cell(upr_end + 1, 2).value
        if starts_with_upr(next_val):
            upr_end += 1
        elif not next_val or str(next_val).strip() == '':
            # Дозволяємо один порожній рядок, якщо далі є УПР
            if upr_end + 1 < MAX_ROW and starts_with_upr(ws.cell(upr_end + 2, 2).value):
                upr_end += 1
            else:
                break
        else:
            break
    
    return upr_end


# ============== ОЧИЩЕННЯ НИЖЧЕ УПР ==============

def clear_below_upr(ws: openpyxl.worksheet.worksheet.Worksheet, upr_end: int):
    """
    Очищає дані B..FP (2..172) нижче блоку УПР, але зберігає формули у H, I, ET.
    """
    first_row = upr_end + 1
    if first_row > MAX_ROW:
        return
    
    # Знайдемо найглибший зайнятий рядок серед B..FP (окрім H/I/ET)
    last_row = first_row
    for col in range(COL_START, COL_END + 1):
        if col in FORMULA_COLS:
            continue
        for row in range(ws.max_row, first_row - 1, -1):
            if ws.cell(row, col).value is not None:
                if row > last_row:
                    last_row = row
                break
    
    if last_row > MAX_ROW:
        last_row = MAX_ROW
    
    # Очищуємо всі колонки окрім формульних
    for row in range(first_row, last_row + 1):
        for col in range(COL_START, COL_END + 1):
            if col not in FORMULA_COLS:
                set_cell_value(ws, row, col, None)


# ============== ЗБІР ДАНИХ З ФАЙЛІВ ==============

def find_excel_files(folder: str) -> List[str]:
    """
    Рекурсивний пошук Excel файлів з шаблонами "Дод.10", "Дод 10", "Додаток 10".
    Ігнорує тимчасові файли (~$...)
    """
    files = []
    folder_path = Path(folder)
    
    if not folder_path.exists():
        return files
    
    # Компілюємо regex шаблони (case-insensitive)
    patterns = [re.compile(p, re.IGNORECASE) for p in TOKEN_PATTERNS]
    
    for path in folder_path.rglob('*'):
        if path.is_file() and not path.name.startswith('~$'):
            # Перевіряємо чи назва файлу відповідає хоча б одному шаблону
            name_lower = path.name.lower()
            if any(pattern.search(name_lower) for pattern in patterns):
                ext = path.suffix.lower()
                if ext in ('.xlsx', '.xlsm', '.xlsb', '.xls'):
                    files.append(str(path))
    
    return files


def row_to_key(row_values: List) -> str:
    """Створює ключ дедуплікації з рядка даних"""
    return chr(30).join(str(v) if v is not None else '' for v in row_values)


def collect_from_file(file_path: str, buckets: Dict[str, List[List]], global_seen: Set[str]):
    """
    Збирає дані з одного файлу-джерела (БЕЗ УПР).
    У файлах-джерелах просто читаємо всі рядки, де в колонці B є назва підрозділу.
    
    - buckets: {unit_name: [[row_values], ...]}
    - global_seen: множина ключів для глобального дедупу
    """
    try:
        # 🚀 ОПТИМІЗАЦІЯ: read_only=True для швидшого читання (~30% прискорення)
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True, keep_vba=False)
    except Exception as e:
        print(f"⚠️ Не вдалося відкрити {os.path.basename(file_path)}: {e}", file=sys.stderr)
        return
    
    local_seen = set()
    collected = 0
    
    # Читаємо ПЕРШИЙ лист (зазвичай там дані)
    if len(wb.worksheets) == 0:
        wb.close()
        return
    
    ws = wb.worksheets[0]
    max_row = ws.max_row or 100
    
    # Обмежуємо кількість рядків для читання
    if max_row > 1000:
        max_row = 1000

    # 🚀 ОПТИМІЗАЦІЯ: Використовуємо iter_rows для швидшого доступу (2-3x швидше)
    # Читаємо всі колонки B:FP (2..172) одразу
    for row_data in ws.iter_rows(min_row=1, max_row=max_row, min_col=COL_START, max_col=COL_END, values_only=True):
        # Перша колонка row_data[0] = колонка B (назва підрозділу)
        unit_raw = row_data[0]
        if not unit_raw:
            continue

        # Нормалізуємо назву підрозділу
        unit_str = str(unit_raw).strip()
        if not unit_str:
            continue

        unit_canon = canon_unit(unit_str)
        if not unit_canon:
            continue

        # row_data вже містить весь рядок B:FP (2..172)
        row_values = list(row_data)
        
        # Перевірка: чи весь рядок порожній (крім назви підрозділу)
        has_data = any(v is not None and str(v).strip() for v in row_values[1:])
        if not has_data:
            continue
        
        # Ключ дедупу: unit + весь рядок
        key = f"{unit_canon}|{row_to_key(row_values)}"
        
        if key not in local_seen and key not in global_seen:
            local_seen.add(key)
            global_seen.add(key)
            
            if unit_canon not in buckets:
                buckets[unit_canon] = []
            buckets[unit_canon].append(row_values)
            collected += 1
    
    wb.close()
    
    # Виводимо скільки зібрали з цього файлу
    if collected > 0:
        print(f"      ✓ Зібрано {collected} унікальних рядків")


# ============== ВСТАВКА У ВИХІДНИЙ ФАЙЛ ==============

def paste_row_values(ws: openpyxl.worksheet.worksheet.Worksheet, target_row: int, values: List):
    """
    Вставляє значення з values у target_row, пропускаючи формульні колонки H, I, ET.

    Mapping:
    - values[0..5]   -> B..G   (col 2..7)
    - values[6..7]   -> H..I   (col 8..9) - ПРОПУСКАЄМО (формули)
    - values[8..147] -> J..ES  (col 10..149)
    - values[148]    -> ET     (col 150) - ПРОПУСКАЄМО (формула)
    - values[149..170] -> EU..FP (col 151..172)
    """
    if target_row > MAX_ROW:
        return

    # B..G (col 2..7) => values[0..5]
    last_unit = getattr(paste_row_values, "_last_unit", "")
    for i, col in enumerate(range(2, 8)):  # 2,3,4,5,6,7
        cell_value = values[i]
        if col == 2:  # Колонка B використовується для перевірок у COM
            if cell_value is None or str(cell_value).strip() == "":
                cell_value = last_unit
            else:
                last_unit = cell_value
                paste_row_values._last_unit = cell_value
        set_cell_value(ws, target_row, col, cell_value)

    # Якщо колонка B досі порожня (даних немає) - очищуємо пам'ять попереднього підрозділу
    if getattr(paste_row_values, "_last_unit", "") and (
        values[0] is None or str(values[0]).strip() == ""
    ) and all(
        (values[idx] is None or str(values[idx]).strip() == "")
        for idx in range(1, len(values))
    ):
        paste_row_values._last_unit = ""

    # Пропускаємо H (col 8) - зберігаємо формулу
    # Пропускаємо I (col 9) - зберігаємо формулу

    # J..ES (col 10..149) => values[8..147]
    for i, col in enumerate(range(10, 150)):  # 10..149 (140 колонок)
        set_cell_value(ws, target_row, col, values[8 + i])

    # Пропускаємо ET (col 150) - зберігаємо формулу

    # EU..FP (col 151..172) => values[149..170]
    for i, col in enumerate(range(151, 173)):  # 151..172 (22 колонки)
        set_cell_value(ws, target_row, col, values[149 + i])


# ============== ВИПРАВЛЕННЯ ЗВАНЬ/ПОСАД ==============

def load_corrections_index(corrections_file: str, value_col: int) -> Dict[str, str]:
    """
    Завантажити індекс виправлень з файлу Excel

    Args:
        corrections_file: Шлях до файлу з виправленнями
        value_col: Колонка зі значенням для виправлення (3=звання, інша=посада)

    Returns:
        Словник {нормалізований_ПІБ: значення}
    """
    from text_utils import normalize_text

    corrections_index = {}

    try:
        # 🚀 ОПТИМІЗАЦІЯ: read_only=True для швидшого читання
        wb = openpyxl.load_workbook(corrections_file, data_only=True, read_only=True)

        # Шукаємо по всіх аркушах
        for sheet in wb.worksheets:
            for row_idx in range(1, sheet.max_row + 1):
                # ПІБ у колонці D (4)
                pib_cell = sheet.cell(row=row_idx, column=4)
                pib = pib_cell.value

                if not pib:
                    continue

                # Нормалізуємо ПІБ (ігноруємо зайві пробіли)
                pib_normalized = normalize_text(str(pib), remove_spaces=False)

                if not pib_normalized:
                    continue

                # Отримуємо значення з потрібної колонки
                value_cell = sheet.cell(row=row_idx, column=value_col)
                value = value_cell.value

                if value:
                    corrections_index[pib_normalized] = str(value).strip()

        wb.close()
        print(f"✅ Завантажено індекс виправлень: {len(corrections_index)} записів")

    except Exception as e:
        print(f"⚠️ Помилка завантаження файлу виправлень: {e}")

    return corrections_index


def apply_corrections(dest_ws, start_row: int, end_row: int, corrections_index: Dict[str, str],
                     target_col: int, col_name: str) -> List[str]:
    """
    Застосувати виправлення до колонки

    Args:
        dest_ws: Аркуш призначення
        start_row: Початковий рядок
        end_row: Кінцевий рядок
        corrections_index: Індекс виправлень {ПІБ: значення}
        target_col: Колонка для виправлення (4=звання, інша=посада)
        col_name: Назва колонки для логування

    Returns:
        Список логів про виправлення
    """
    from text_utils import normalize_text

    correction_logs = []
    corrections_count = 0

    # ПІБ у колонці E (5)
    PIB_COL = 5

    for row_idx in range(start_row, end_row + 1):
        # Отримуємо ПІБ
        pib_cell = dest_ws.cell(row=row_idx, column=PIB_COL)
        pib = pib_cell.value

        if not pib:
            continue

        # Нормалізуємо ПІБ
        pib_normalized = normalize_text(str(pib), remove_spaces=False)

        if not pib_normalized:
            continue

        # Шукаємо в індексі виправлень
        if pib_normalized in corrections_index:
            new_value = corrections_index[pib_normalized]

            # Отримуємо поточне значення
            current_cell = get_writable_cell(dest_ws, row_idx, target_col)
            current_value = str(current_cell.value).strip() if current_cell.value else ""

            # Порівнюємо
            if current_value != new_value:
                # Виправляємо
                set_cell_value(dest_ws, row_idx, target_col, new_value)
                corrections_count += 1

                log_entry = f"   ✍️ Рядок {row_idx} | {pib} | {col_name}: «{current_value}» → «{new_value}»"
                correction_logs.append(log_entry)

    if corrections_count > 0:
        summary = f"✅ Виправлено {col_name}: {corrections_count} записів"
        print(summary)
        correction_logs.insert(0, summary)
    else:
        summary = f"ℹ️ Виправлень {col_name} не знайдено"
        print(summary)
        correction_logs.insert(0, summary)

    return correction_logs


# ============== ГОЛОВНА ФУНКЦІЯ ==============

def process_dodatok10(
    input_folder: str,
    destination_file: str,
    auto_open: bool = False,
    ignore_formula_cols: bool = True
):
    import datetime
    try:
        # 1. Перевірка існування папки
        if not os.path.exists(input_folder):
            return {"ok": False, "error": f"Папка не існує: {input_folder}"}
        # 2. Перевірка існування цільового файлу
        if not os.path.exists(destination_file):
            return {"ok": False, "error": f"Цільовий файл не існує: {destination_file}"}
        # 3. Пошук файлів
        print(f"🔍 Пошук файлів з шаблонами 'Дод.10', 'Дод 10', 'Додаток 10' у {input_folder}...")
        files = find_excel_files(input_folder)
        if not files:
            return {"ok": False, "error": "Не знайдено жодного файлу з назвою 'Дод.10', 'Дод 10' або 'Додаток 10'"}
        print(f"✅ Знайдено файлів: {len(files)}")
        # 4. Відкриття цільового файлу
        print(f"📂 Відкриття цільового файлу: {destination_file}")
        dest_wb = openpyxl.load_workbook(destination_file)
        # 5. Знаходження аркуша з УПР
        dest_ws = detect_target_sheet(dest_wb)
        if not dest_ws:
            dest_wb.close()
            return {"ok": False, "error": "Не знайдено аркуша з блоком 'УПР' у колонці B"}
        print(f"✅ Знайдено аркуш: {dest_ws.title}")
        # 6. Знаходження кінця блоку УПР
        upr_end = find_upr_end(dest_ws)
        if upr_end == 0:
            dest_wb.close()
            return {"ok": False, "error": "Не вдалося визначити кінець блоку 'УПР'"}
        print(f"✅ Кінець УПР: рядок {upr_end}")
        # 7. Очищення нижче УПР (зберігаючи формули H/I/ET)
        if ignore_formula_cols:
            print(f"🧹 Очищення даних нижче УПР (зберігаючи формули H/I/ET)...")
            clear_below_upr(dest_ws, upr_end)
        # 8. Збір даних з усіх файлів
        print(f"📊 Збір даних з {len(files)} файлів...")
        buckets = {}
        global_seen = set()
        logs = []
        for idx, file_path in enumerate(files, 1):
            file_name = os.path.basename(file_path)
            log_line = f"[{idx}/{len(files)}] {file_name}"
            print(f"  {log_line}")
            logs.append(log_line)
            # Збір даних з файлу
            before = sum(len(rows) for rows in buckets.values())
            collect_from_file(file_path, buckets, global_seen)
            after = sum(len(rows) for rows in buckets.values())
            logs.append(f"      ✓ Зібрано {after - before} унікальних рядків")
        # 9. Підрахунок зібраних рядків
        total_rows = sum(len(rows) for rows in buckets.values())
        print(f"✅ Зібрано унікальних рядків: {total_rows}")
        logs.append(f"✅ Зібрано унікальних рядків: {total_rows}")
        # 10. Запис у фіксованому порядку підрозділів
        write_row = upr_end + 1
        capacity = MAX_ROW - write_row + 1
        written = 0
        if capacity <= 0:
            dest_wb.close()
            return {"ok": False, "error": f"Немає місця нижче УПР. Область запису закінчується до рядка {MAX_ROW}"}
        print(f"✍️ Запис даних (межа: {MAX_ROW})...")
        logs.append(f"✍️ Запис даних (межа: {MAX_ROW})...")
        setattr(paste_row_values, "_last_unit", "")
        for unit in UNITS_ORDER:
            if unit not in buckets:
                continue
            rows = buckets[unit]
            unit_log = f"  • {unit}: {len(rows)} рядків"
            print(unit_log)
            logs.append(unit_log)
            for row_values in rows:
                if written >= capacity:
                    cut_log = f"⚠️ Дані обрізано: досягнуто межі {MAX_ROW}. Записано {written} рядків."
                    print(cut_log)
                    logs.append(cut_log)
                    break
                paste_row_values(dest_ws, write_row, row_values)
                write_row += 1
                written += 1
            if written >= capacity:
                break

        # 11. Додаємо лист LOG з логами
        if 'LOG' in dest_wb.sheetnames:
            del dest_wb['LOG']
        log_ws = dest_wb.create_sheet('LOG', 0)  # Створюємо на першому місці

        from openpyxl.styles import Font, PatternFill, Alignment

        # === ЗАГОЛОВОК ===
        log_ws['A1'] = '📋 ЛОГ ОБРОБКИ ДОДАТКУ 10'
        log_ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        log_ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        log_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        log_ws.row_dimensions[1].height = 30

        # Підзаголовок
        now = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        log_ws['A2'] = f'Дата обробки: {now}'
        log_ws['A2'].font = Font(size=10, italic=True, color='7F7F7F')
        log_ws['A2'].alignment = Alignment(horizontal='center')

        log_ws['A3'] = ''

        # === СТАТИСТИКА ===
        log_ws['A4'] = '📊 ПІДСУМОК ОБРОБКИ'
        log_ws['A4'].font = Font(bold=True, size=12, color='FFFFFF')
        log_ws['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        log_ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
        log_ws.row_dimensions[4].height = 25

        stats_data = [
            ('📁 Оброблено файлів:', len(files)),
            ('📝 Записано рядків:', written),
            ('🎯 Знайдено підрозділів:', len(buckets)),
            ('📏 Доступна ємність:', capacity),
            ('⚙️ Кінець блоку УПР:', f'рядок {upr_end}')
        ]

        row = 5
        for label, value in stats_data:
            log_ws.cell(row, 1, label)
            log_ws.cell(row, 2, value)
            log_ws.cell(row, 1).font = Font(bold=True, size=10)
            log_ws.cell(row, 2).font = Font(size=10, color='0070C0')
            log_ws.cell(row, 2).alignment = Alignment(horizontal='right')
            row += 1

        log_ws.cell(row, 1, '')
        row += 1

        # === ДЕТАЛЬНИЙ ЛОГ ===
        log_ws.cell(row, 1, '📝 ДЕТАЛЬНИЙ ЛОГ ПРОЦЕСУ')
        log_ws.cell(row, 1).font = Font(bold=True, size=12, color='FFFFFF')
        log_ws.cell(row, 1).fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        log_ws.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')
        log_ws.row_dimensions[row].height = 25
        detail_row = row
        row += 1

        # Стилі для логів
        for log_line in logs:
            cell = log_ws.cell(row, 1, log_line)

            # Кольори за емодзі
            if log_line.strip().startswith('✅'):
                cell.font = Font(bold=True, color='228B22')
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            elif log_line.strip().startswith('⚠️'):
                cell.font = Font(bold=True, color='FF8C00')
                cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            elif log_line.strip().startswith('❌'):
                cell.font = Font(bold=True, color='C00000')
                cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            elif log_line.strip().startswith('✍️'):
                cell.font = Font(color='4472C4')
            elif log_line.strip().startswith('✓'):
                cell.font = Font(color='70AD47')
            elif log_line.strip().startswith('🔍'):
                cell.font = Font(color='7030A0')
            elif log_line.strip().startswith('📂'):
                cell.font = Font(color='203764')
            elif log_line.strip().startswith('📊'):
                cell.font = Font(color='0070C0')
            elif log_line.strip().startswith('🧹'):
                cell.font = Font(color='C55A11')
            elif log_line.strip().startswith('💾'):
                cell.font = Font(color='44546A')
            elif log_line.startswith('  '):
                cell.font = Font(size=9, color='595959')
                cell.alignment = Alignment(indent=1)
            else:
                cell.font = Font(color='404040')

            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            row += 1

        # Налаштування колонок
        log_ws.column_dimensions['A'].width = 100
        log_ws.column_dimensions['B'].width = 20

        # Об'єднання комірок для заголовків
        log_ws.merge_cells('A1:B1')
        log_ws.merge_cells('A2:B2')
        log_ws.merge_cells('A4:B4')
        log_ws.merge_cells(f'A{detail_row}:B{detail_row}')

        # === FNP ПЕРЕВІРКА (через Excel COM API) ===
        fnp_errors = []
        # Зберігаємо назву аркуша (може знадобитися для FNP або Дублів)
        target_sheet_name = dest_ws.title

        fnp_env = os.environ.get('PY_FNP_CHECK', '0')
        if fnp_env == '1':
            print(f"🔍 FNP перевірка колонки H (рядки 10-900) через Excel COM...")

            # Спочатку зберігаємо файл, щоб Excel міг його відкрити
            print(f"💾 Збереження файлу перед FNP перевіркою...")
            dest_wb.save(destination_file)
            dest_wb.close()

            try:
                import win32com.client

                # Відкриваємо Excel через COM
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False

                # Відкриваємо файл
                wb_com = excel.Workbooks.Open(os.path.abspath(destination_file))

                # Знаходимо аркуш
                ws_com = wb_com.Worksheets(target_sheet_name)

                checked_count = 0

                # Перевіряємо рядки 10-900
                for check_row in range(10, MAX_ROW + 1):
                    # Перевіряємо чи є підрозділ у колонці B
                    unit_val = ws_com.Cells(check_row, 2).Value
                    if not unit_val or not str(unit_val).strip():
                        continue

                    unit_name = str(unit_val).strip()

                    # Читаємо ОБЧИСЛЕНЕ значення H через COM
                    h_val = ws_com.Cells(check_row, 8).Value
                    checked_count += 1

                    # Перевіряємо значення H
                    if h_val is None:
                        msg = f"Рядок {check_row} — Підрозділ: «{unit_name}», ФнП(H)=порожньо"
                        fnp_errors.append(msg)
                    else:
                        try:
                            h_num = float(h_val)
                            if h_num != 1.0:
                                msg = f"Рядок {check_row} — Підрозділ: «{unit_name}», ФнП(H)={h_val}"
                                fnp_errors.append(msg)
                        except (ValueError, TypeError):
                            msg = f"Рядок {check_row} — Підрозділ: «{unit_name}», ФнП(H)={h_val} (некоректне)"
                            fnp_errors.append(msg)

                # Закриваємо Excel
                wb_com.Close(SaveChanges=False)
                excel.Quit()

                print(f"   Перевірено рядків: {checked_count}, знайдено помилок: {len(fnp_errors)}")

                # Відкриваємо файл знову через openpyxl для додавання FNP логів
                dest_wb = openpyxl.load_workbook(destination_file)
                dest_ws = dest_wb[target_sheet_name]
                log_ws = dest_wb['LOG']

                # Знаходимо останній заповнений рядок у LOG
                row = log_ws.max_row

            except ImportError:
                print(f"⚠️ win32com не встановлено. FNP перевірка пропущена.")
                print(f"   Встановіть: pip install pywin32")
                # Відкриваємо файл знову
                dest_wb = openpyxl.load_workbook(destination_file)
                dest_ws = dest_wb[target_sheet_name]
                log_ws = dest_wb['LOG']
                row = log_ws.max_row
            except Exception as e:
                print(f"⚠️ Помилка FNP перевірки через COM: {e}")
                import traceback
                traceback.print_exc()
                # Відкриваємо файл знову
                dest_wb = openpyxl.load_workbook(destination_file)
                dest_ws = dest_wb[target_sheet_name]
                log_ws = dest_wb['LOG']
                row = log_ws.max_row

        # Додаємо секцію FNP до LOG (тільки якщо FNP перевірка була увімкнена)
        if fnp_env == '1':
            if fnp_errors:
                row += 1
                log_ws.cell(row, 1, '')
                row += 1
                log_ws.cell(row, 1, '⚠️ FNP ПОМИЛКИ')
                log_ws.cell(row, 1).font = Font(bold=True, size=12, color='FFFFFF')
                log_ws.cell(row, 1).fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
                log_ws.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')
                log_ws.row_dimensions[row].height = 25
                fnp_header_row = row
                row += 1

                log_ws.cell(row, 1, f'Знайдено помилок: {len(fnp_errors)}')
                log_ws.cell(row, 1).font = Font(bold=True, color='FF8C00')
                row += 1

                # Додаємо перші 50 помилок
                for i, err in enumerate(fnp_errors[:50], 1):
                    cell = log_ws.cell(row, 1, f"{i}. {err}")
                    cell.font = Font(size=9, color='C00000')
                    cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    row += 1

                if len(fnp_errors) > 50:
                    cell = log_ws.cell(row, 1, f"... і ще {len(fnp_errors) - 50} помилок")
                    cell.font = Font(italic=True, color='7F7F7F')
                    row += 1

                log_ws.merge_cells(f'A{fnp_header_row}:B{fnp_header_row}')
                print(f"⚠️ FNP: Знайдено помилок: {len(fnp_errors)}")
            else:
                row += 1
                log_ws.cell(row, 1, '')
                row += 1
                log_ws.cell(row, 1, '✅ FNP ПЕРЕВІРКА: Всі рядки правильні')
                log_ws.cell(row, 1).font = Font(bold=True, color='228B22')
                log_ws.cell(row, 1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                log_ws.merge_cells(f'A{row}:B{row}')
                print(f"✅ FNP: Помилок не знайдено")

        # === ПЕРЕВІРКА ДУБЛІВ (через Excel COM API) ===
        duplicates_errors = []
        duplicates_env = os.environ.get('PY_DUPLICATES_CHECK', '0')
        if duplicates_env == '1':
            print(f"🔍 Перевірка дублів ПІБ у колонці E (рядки 10-900) через Excel COM...")

            # Якщо dest_wb ще відкритий - зберігаємо та закриваємо
            try:
                dest_wb.save(destination_file)
                dest_wb.close()
            except:
                pass

            try:
                import win32com.client
                from text_utils import normalize_text

                # Відкриваємо Excel через COM
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False

                # Відкриваємо файл
                wb_com = excel.Workbooks.Open(os.path.abspath(destination_file))

                # Знаходимо аркуш
                ws_com = wb_com.Worksheets(target_sheet_name)

                # Словник для відстеження дублів: {нормалізований_ПІБ: [(рядок, підрозділ, оригінальний_ПІБ), ...]}
                seen_values = {}
                checked_count = 0

                # Перевіряємо рядки 10-900
                for check_row in range(10, MAX_ROW + 1):
                    # Перевіряємо чи є підрозділ у колонці B
                    unit_val = ws_com.Cells(check_row, 2).Value
                    if not unit_val or not str(unit_val).strip():
                        continue

                    unit_name = str(unit_val).strip()

                    # Читаємо значення з колонки E (ПІБ)
                    pib_val = ws_com.Cells(check_row, 5).Value
                    checked_count += 1

                    # Якщо значення порожнє - пропускаємо
                    if pib_val is None or str(pib_val).strip() == '':
                        continue

                    pib_original = str(pib_val).strip()

                    # Нормалізуємо ПІБ: нижній регістр, видалення зайвих пробілів, латинські→кирилічні
                    # remove_spaces=False щоб зберегти пробіли між словами
                    pib_normalized = normalize_text(pib_original, remove_spaces=False, case='lower')

                    # Додаємо до словника
                    if pib_normalized not in seen_values:
                        seen_values[pib_normalized] = []
                    seen_values[pib_normalized].append((check_row, unit_name, pib_original))

                # Закриваємо Excel
                wb_com.Close(SaveChanges=False)
                excel.Quit()

                # Знаходимо дублі (значення які зустрічаються більше 1 разу)
                for pib_norm, rows_list in seen_values.items():
                    if len(rows_list) > 1:
                        # Беремо оригінальне написання з першого входження
                        first_original = rows_list[0][2]
                        # Формуємо повідомлення про дублі
                        rows_info = ', '.join([f"рядок {r} ({unit})" for r, unit, _ in rows_list])
                        msg = f"ПІБ «{first_original}» повторюється {len(rows_list)} разів: {rows_info}"
                        duplicates_errors.append(msg)

                print(f"   Перевірено рядків: {checked_count}, знайдено дублів: {len(duplicates_errors)}")

                # Відкриваємо файл знову через openpyxl для додавання логів
                dest_wb = openpyxl.load_workbook(destination_file)
                dest_ws = dest_wb[target_sheet_name]
                log_ws = dest_wb['LOG']

                # Знаходимо останній заповнений рядок у LOG
                row = log_ws.max_row

            except ImportError:
                print(f"⚠️ win32com не встановлено. Перевірка дублів пропущена.")
                print(f"   Встановіть: pip install pywin32")
                # Відкриваємо файл знову
                dest_wb = openpyxl.load_workbook(destination_file)
                dest_ws = dest_wb[target_sheet_name]
                log_ws = dest_wb['LOG']
                row = log_ws.max_row
            except Exception as e:
                print(f"⚠️ Помилка перевірки дублів через COM: {e}")
                import traceback
                traceback.print_exc()
                # Відкриваємо файл знову
                dest_wb = openpyxl.load_workbook(destination_file)
                dest_ws = dest_wb[target_sheet_name]
                log_ws = dest_wb['LOG']
                row = log_ws.max_row

        # Додаємо секцію ДУБЛІ до LOG (тільки якщо перевірка була увімкнена)
        if duplicates_env == '1':
            if duplicates_errors:
                row += 1
                log_ws.cell(row, 1, '')
                row += 1
                log_ws.cell(row, 1, '⚠️ ЗНАЙДЕНО ДУБЛІ')
                log_ws.cell(row, 1).font = Font(bold=True, size=12, color='FFFFFF')
                log_ws.cell(row, 1).fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
                log_ws.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')
                log_ws.row_dimensions[row].height = 25
                duplicates_header_row = row
                row += 1

                log_ws.cell(row, 1, f'Знайдено дублів: {len(duplicates_errors)}')
                log_ws.cell(row, 1).font = Font(bold=True, color='FF8C00')
                row += 1

                # Додаємо перші 50 дублів
                for i, err in enumerate(duplicates_errors[:50], 1):
                    cell = log_ws.cell(row, 1, f"{i}. {err}")
                    cell.font = Font(size=9, color='C00000')
                    cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    row += 1

                if len(duplicates_errors) > 50:
                    cell = log_ws.cell(row, 1, f"... і ще {len(duplicates_errors) - 50} дублів")
                    cell.font = Font(italic=True, color='7F7F7F')
                    row += 1

                log_ws.merge_cells(f'A{duplicates_header_row}:B{duplicates_header_row}')
                print(f"⚠️ ДУБЛІ: Знайдено дублів: {len(duplicates_errors)}")
            else:
                row += 1
                log_ws.cell(row, 1, '')
                row += 1
                log_ws.cell(row, 1, '✅ ПЕРЕВІРКА ДУБЛІВ: Дублів не знайдено')
                log_ws.cell(row, 1).font = Font(bold=True, color='228B22')
                log_ws.cell(row, 1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                log_ws.merge_cells(f'A{row}:B{row}')
                print(f"✅ ДУБЛІ: Дублів не знайдено")

        # === ПЕРЕВІРКА СТРОЙОВКИ (через Excel COM API) ===
        stroiovka_errors = []
        stroiovka_env = os.environ.get('PY_STROIOVKA_CHECK', '0')
        stroiovka_file_path = os.environ.get('PY_STROIOVKA_FILE', '').strip().strip('"').strip("'")

        if stroiovka_env == '1' and stroiovka_file_path:
            print(f"🔍 Перевірка відповідності зі стройовкою через Excel COM...")
            print(f"📊 Файл стройовки: {stroiovka_file_path}")

            # Перевіряємо чи існує файл стройовки
            if not os.path.exists(stroiovka_file_path):
                error_msg = f"❌ Файл стройовки не знайдено: {stroiovka_file_path}"
                print(error_msg)
                logs.append(error_msg)
                # Продовжуємо без перевірки стройовки
            else:
                # Якщо dest_wb ще відкритий - зберігаємо та закриваємо
                try:
                    dest_wb.save(destination_file)
                    dest_wb.close()
                except Exception as e:
                    print(f"⚠️ Попередження при збереженні перед перевіркою стройовки: {e}")

                try:
                    import win32com.client

                    # Відкриваємо Excel через COM
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False

                    # Відкриваємо обидва файли
                    # UpdateLinks=0 - не оновлювати зовнішні посилання
                    # ReadOnly=True - відкрити тільки для читання
                    # CorruptLoad=2 - спробувати відновити пошкоджений файл
                    try:
                        dest_wb_com = excel.Workbooks.Open(
                            os.path.abspath(destination_file),
                            UpdateLinks=0,
                            ReadOnly=False,
                            CorruptLoad=2
                        )
                    except Exception as e:
                        excel.Quit()
                        raise Exception(f"Не вдалося відкрити файл результату. Можливо, він відкритий в Excel. Закрийте файл та спробуйте знову. Помилка: {e}")

                    try:
                        stroiovka_wb_com = excel.Workbooks.Open(
                            os.path.abspath(stroiovka_file_path),
                            UpdateLinks=0,
                            ReadOnly=True,
                            CorruptLoad=2
                        )
                    except Exception as e:
                        dest_wb_com.Close(SaveChanges=False)
                        excel.Quit()
                        raise Exception(f"Не вдалося відкрити файл стройовки. Можливо, він відкритий в Excel або не існує. Помилка: {e}")

                    # Знаходимо аркуші
                    dest_ws_com = dest_wb_com.Worksheets(target_sheet_name)
                    stroiovka_ws_com = stroiovka_wb_com.Worksheets('ЗС')  # Лист "ЗС" стройовки

                    # Створюємо словник ПІБ -> статус зі стройовки
                    stroiovka_data = {}  # {ПІБ: статус}

                    # Імпортуємо функцію нормалізації
                    from text_utils import normalize_text

                    # Читаємо стройовку (колонка D - ПІБ, колонка F - статус)
                    for check_row in range(1, 1201):
                        pib = stroiovka_ws_com.Cells(check_row, 4).Value  # Колонка D
                        status = stroiovka_ws_com.Cells(check_row, 6).Value  # Колонка F

                        # Пропускаємо порожні рядки та рядки без ПІБ
                        if not pib or not str(pib).strip():
                            continue
                        
                        # ВАЖЛИВО: Пропускаємо рядки без статусу (заголовки, підрозділи)
                        if not status or not str(status).strip():
                            continue

                        # Нормалізуємо ПІБ: латинські→кирилічні, NBSP→пробіл, upper регістр
                        pib_str = normalize_text(pib, remove_spaces=False, case='upper')
                        # Нормалізуємо статус до верхнього регістру для уніфікації
                        status_str = normalize_text(status, remove_spaces=False, case='upper')
                        stroiovka_data[pib_str] = status_str

                    print(f"   Завантажено {len(stroiovka_data)} записів зі стройовки")
                    
                    # Діагностика: виводимо перші 5 ПІБ зі стройовки
                    sample_pibs = list(stroiovka_data.keys())[:5]
                    print(f"   Приклади ПІБ зі стройовки: {sample_pibs}")
                    
                    # Діагностика: виводимо унікальні статуси
                    unique_statuses = set(stroiovka_data.values())
                    print(f"   Унікальні статуси ({len(unique_statuses)}): {sorted(unique_statuses)}")

                    checked_count = 0

                    # Перевіряємо файл результату (рядки 10-900)
                    for check_row in range(10, MAX_ROW + 1):
                        try:
                            # Перевіряємо чи є підрозділ у колонці B
                            unit_val = dest_ws_com.Cells(check_row, 2).Value
                            if not unit_val or not str(unit_val).strip():
                                continue

                            unit_name = str(unit_val).strip()

                            # Читаємо ПІБ з колонки E та значення I
                            pib_val = dest_ws_com.Cells(check_row, 5).Value  # Колонка E
                            i_val = dest_ws_com.Cells(check_row, 9).Value   # Колонка I
                            g_val = dest_ws_com.Cells(check_row, 7).Value   # Колонка G (МВ/ін)
                            es_val = dest_ws_com.Cells(check_row, 149).Value  # Колонка ES (прапор БЗ з EU:FJ)
                            eu_val = dest_ws_com.Cells(check_row, 151).Value  # Колонка EU (прапор ВЛК)
                            ev_val = dest_ws_com.Cells(check_row, 152).Value  # Колонка EV (прапор Ш)
                            ew_val = dest_ws_com.Cells(check_row, 153).Value  # Колонка EW (прапор Ш)
                            ex_val = dest_ws_com.Cells(check_row, 154).Value  # Колонка EX (прапор В+МВ)
                            ey_val = dest_ws_com.Cells(check_row, 155).Value  # Колонка EY (прапор В без МВ)
                            fc_val = dest_ws_com.Cells(check_row, 159).Value  # Колонка FC (прапор ВД)
                            fe_val = dest_ws_com.Cells(check_row, 161).Value  # Колонка FE (прапор БЗН)
                            ff_val = dest_ws_com.Cells(check_row, 162).Value  # Колонка FF (прапор СЗЧ)
                            fj_val = dest_ws_com.Cells(check_row, 166).Value  # Колонка FJ (прапор РБ)
                            fn_val = dest_ws_com.Cells(check_row, 170).Value  # Колонка FN (прапор КЗВ/СП/КР/БЧ/РАО)
                            
                            # Читаємо всі колонки EU:FJ (151-166) для перевірки БЗ
                            flags_eu_fn = []
                            for col in range(151, 167):  # EU(151) до FJ(166) включно
                                val = dest_ws_com.Cells(check_row, col).Value
                                flags_eu_fn.append(val)

                            if not pib_val or not str(pib_val).strip():
                                continue

                            # Нормалізуємо ПІБ: латинські→кирилічні, NBSP→пробіл, upper регістр
                            pib_str = normalize_text(pib_val, remove_spaces=False, case='upper')
                        
                        except Exception as row_error:
                            # Якщо помилка COM на конкретному рядку, пропускаємо його та продовжуємо
                            print(f"   ⚠️ Помилка при читанні рядка {check_row}: {row_error}")
                            continue

                        # Допоміжна функція для перевірки "== 1"
                        def is_one(value) -> bool:
                            if value is None:
                                return False
                            try:
                                return float(value) == 1.0
                            except (ValueError, TypeError):
                                return str(value).strip() == '1'

                        # Перевіряємо чи колонка I = 1
                        try:
                            i_num = float(i_val) if i_val is not None else 0
                        except:
                            i_num = 0

                        if i_num == 1.0:
                            checked_count += 1
                            # Перевіряємо наявність у стройовці
                            if pib_str in stroiovka_data:
                                stroiovka_status = stroiovka_data[pib_str]
                                # Перевіряємо чи статус = "БЗ"
                                if stroiovka_status != 'БЗ':
                                    msg = f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}», Статус у стройовці: «{stroiovka_status}» (очікується «БЗ»)"
                                    stroiovka_errors.append(msg)
                            else:
                                # Діагностика: для проблемних підрозділів виводимо додатково
                                if unit_name in ["Взвод зв'язку", "РМТЗ", "Медичний пункт"]:
                                    print(f"   ⚠️ Не знайдено: підрозділ='{unit_name}', ПІБ='{pib_str}' (довжина={len(pib_str)})")
                                    # Шукаємо схожі ПІБ у стройовці
                                    similar = [p for p in stroiovka_data.keys() if pib_str[:10] in p or p[:10] in pib_str]
                                    if similar:
                                        print(f"      Схожі в стройовці: {similar[:3]}")
                                
                                msg = f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}» — не знайдено у стройовці"
                                stroiovka_errors.append(msg)

                        # АВТОМАТИЧНЕ ПРОСТАВЛЯННЯ ПРАПОРЦІВ НА ОСНОВІ СТАТУСУ
                        # Якщо особа є у стройовці, автоматично проставляємо відповідні прапорці
                        try:
                            if pib_str in stroiovka_data:
                                stroiovka_status = stroiovka_data[pib_str]
                                
                                # Відповідність статус → колонка:
                                # ВЛК → EU (151)
                                # Ш → EV (152) та EW (153)
                                # В → EX (154) та EY (155)
                                
                                if stroiovka_status == 'ВЛК':
                                    # Ставимо EU=1, решту скидаємо
                                    if not is_one(eu_val):
                                        dest_ws_com.Cells(check_row, 151).Value = 1
                                        print(f"   ✓ Рядок {check_row}: автоматично встановлено EU=1 (ВЛК)")
                                    
                                    # Скидаємо ВСІ інші прапорці: J:ES (10-149) + EV:FN (152-166)
                                    cleared = []
                                    # J:ES (10-149)
                                    for col in range(10, 150):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # EV:FN (152-166)
                                    for col in range(152, 167):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    if cleared:
                                        print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус ВЛК)")
                                
                                elif stroiovka_status == 'Ш':
                                    # Ставимо тільки EW=1, решту скидаємо
                                    if not is_one(ew_val):
                                        dest_ws_com.Cells(check_row, 153).Value = 1
                                        print(f"   ✓ Рядок {check_row}: автоматично встановлено EW=1 (Ш)")
                                    
                                    # Скидаємо ВСІ інші прапорці: J:ES (10-149) + EU,EV (151-152) + EX:FN (154-166)
                                    cleared = []
                                    # J:ES (10-149)
                                    for col in range(10, 150):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # EU, EV (151-152)
                                    for col in [151, 152]:
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # EX:FN (154-166)
                                    for col in range(154, 167):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    if cleared:
                                        print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус Ш)")
                                
                                elif stroiovka_status == 'В' or stroiovka_status.startswith("В ("):
                                    # Статус "В" або "В ('МВ)" або "В (...)"
                                    # Перевіряємо значення в колонці G (МВ чи ні)
                                    g_str = str(g_val).strip().upper() if g_val else ''
                                    is_mv = (g_str == 'МВ')
                                    
                                    if is_mv:
                                        # Якщо G = МВ → ставимо EX=1
                                        if not is_one(ex_val):
                                            dest_ws_com.Cells(check_row, 154).Value = 1
                                            print(f"   ✓ Рядок {check_row}: автоматично встановлено EX=1 (В+МВ)")
                                        
                                        # Скидаємо ВСІ інші: J:ES (10-149) + EU:EW (151-153) + EY:FN (155-166)
                                        cleared = []
                                        # J:ES (10-149)
                                        for col in range(10, 150):
                                            val = dest_ws_com.Cells(check_row, col).Value
                                            if is_one(val):
                                                dest_ws_com.Cells(check_row, col).Value = None
                                                cleared.append(str(col))
                                        # EU:EW (151-153)
                                        for col in range(151, 154):
                                            val = dest_ws_com.Cells(check_row, col).Value
                                            if is_one(val):
                                                dest_ws_com.Cells(check_row, col).Value = None
                                                cleared.append(str(col))
                                        # EY:FN (155-166)
                                        for col in range(155, 167):
                                            val = dest_ws_com.Cells(check_row, col).Value
                                            if is_one(val):
                                                dest_ws_com.Cells(check_row, col).Value = None
                                                cleared.append(str(col))
                                        if cleared:
                                            print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус В+МВ)")
                                    else:
                                        # Якщо G ≠ МВ → ставимо EY=1
                                        if not is_one(ey_val):
                                            dest_ws_com.Cells(check_row, 155).Value = 1
                                            print(f"   ✓ Рядок {check_row}: автоматично встановлено EY=1 (В без МВ, G={g_str or 'порожньо'})")
                                        
                                        # Скидаємо ВСІ інші: J:ES (10-149) + EU:EX (151-154) + EZ:FN (156-166)
                                        cleared = []
                                        # J:ES (10-149)
                                        for col in range(10, 150):
                                            val = dest_ws_com.Cells(check_row, col).Value
                                            if is_one(val):
                                                dest_ws_com.Cells(check_row, col).Value = None
                                                cleared.append(str(col))
                                        # EU:EX (151-154)
                                        for col in range(151, 155):
                                            val = dest_ws_com.Cells(check_row, col).Value
                                            if is_one(val):
                                                dest_ws_com.Cells(check_row, col).Value = None
                                                cleared.append(str(col))
                                        # EZ:FN (156-166)
                                        for col in range(156, 167):
                                            val = dest_ws_com.Cells(check_row, col).Value
                                            if is_one(val):
                                                dest_ws_com.Cells(check_row, col).Value = None
                                                cleared.append(str(col))
                                        if cleared:
                                            print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус В без МВ)")
                                
                                elif stroiovka_status == 'БКБП' or stroiovka_status.startswith('БКБП'):
                                    # Статус БКБП → ставимо EZ=1
                                    ez_val = dest_ws_com.Cells(check_row, 156).Value
                                    if not is_one(ez_val):
                                        dest_ws_com.Cells(check_row, 156).Value = 1
                                        print(f"   ✓ Рядок {check_row}: автоматично встановлено EZ=1 (БКБП)")
                                    
                                    # Скидаємо ВСІ інші: J:ES (10-149) + EU:EY (151-155) + FA:FN (157-170)
                                    cleared = []
                                    # J:ES (10-149)
                                    for col in range(10, 150):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # EU:EY (151-155)
                                    for col in range(151, 156):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # FA:FN (157-170)
                                    for col in range(157, 171):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    if cleared:
                                        print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус БКБП, залишено тільки EZ)")
                                
                                elif stroiovka_status == 'ВД':
                                    # Статус ВД → ставимо FC=1 (колонка 159)
                                    fc_val = dest_ws_com.Cells(check_row, 159).Value
                                    if not is_one(fc_val):
                                        dest_ws_com.Cells(check_row, 159).Value = 1
                                        print(f"   ✓ Рядок {check_row}: автоматично встановлено FC=1 (ВД)")
                                    
                                    # Скидаємо ВСІ інші: J:ES (10-149) + EU:FB (151-158) + FD:FN (160-170)
                                    cleared = []
                                    # J:ES (10-149)
                                    for col in range(10, 150):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # EU:FB (151-158)
                                    for col in range(151, 159):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # FD:FN (160-170)
                                    for col in range(160, 171):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    if cleared:
                                        print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус ВД, залишено тільки FC)")
                                
                                elif stroiovka_status in ['КЗВ', 'СП', 'КР', 'БЧ', 'РАО', 'ДИС']:
                                    # Статуси КЗВ, СП, КР, БЧ, РАО, ДИС → ставимо FN=1 (колонка 170)
                                    fn_val = dest_ws_com.Cells(check_row, 170).Value
                                    if not is_one(fn_val):
                                        dest_ws_com.Cells(check_row, 170).Value = 1
                                        print(f"   ✓ Рядок {check_row}: автоматично встановлено FN=1 ({stroiovka_status})")
                                    
                                    # Скидаємо ВСІ інші: J:ES (10-149) + EU:FM (151-169)
                                    cleared = []
                                    # J:ES (10-149)
                                    for col in range(10, 150):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # EU:FM (151-169)
                                    for col in range(151, 170):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    if cleared:
                                        print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус {stroiovka_status}, залишено тільки FN)")
                                
                                elif stroiovka_status == 'РБ':
                                    # Статус РБ → ставимо FJ=1 (колонка 166)
                                    fj_val = dest_ws_com.Cells(check_row, 166).Value
                                    if not is_one(fj_val):
                                        dest_ws_com.Cells(check_row, 166).Value = 1
                                        print(f"   ✓ Рядок {check_row}: автоматично встановлено FJ=1 (РБ)")
                                    
                                    # Скидаємо ВСІ інші: J:ES (10-149) + EU:FI (151-165) + FK:FN (167-170)
                                    cleared = []
                                    # J:ES (10-149)
                                    for col in range(10, 150):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # EU:FI (151-165)
                                    for col in range(151, 166):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # FK:FN (167-170)
                                    for col in range(167, 171):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    if cleared:
                                        print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус РБ, залишено тільки FJ)")
                                
                                elif stroiovka_status == 'БЗН':
                                    # Статус БЗН → ставимо FE=1 (колонка 161)
                                    fe_val = dest_ws_com.Cells(check_row, 161).Value
                                    if not is_one(fe_val):
                                        dest_ws_com.Cells(check_row, 161).Value = 1
                                        print(f"   ✓ Рядок {check_row}: автоматично встановлено FE=1 (БЗН)")
                                    
                                    # Скидаємо ВСІ інші: J:ES (10-149) + EU:FD (151-160) + FF:FJ (162-166)
                                    cleared = []
                                    # J:ES (10-149)
                                    for col in range(10, 150):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # EU:FD (151-160)
                                    for col in range(151, 161):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # FF:FN (162-170)
                                    for col in range(162, 171):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    if cleared:
                                        print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус БЗН, залишено тільки FE)")
                                
                                elif stroiovka_status == 'СЗЧ':
                                    # Статус СЗЧ → ставимо FF=1 (колонка 162)
                                    ff_val = dest_ws_com.Cells(check_row, 162).Value
                                    if not is_one(ff_val):
                                        dest_ws_com.Cells(check_row, 162).Value = 1
                                        print(f"   ✓ Рядок {check_row}: автоматично встановлено FF=1 (СЗЧ)")
                                    
                                    # Скидаємо ВСІ інші: J:ES (10-149) + EU:FE (151-161) + FG:FJ (163-166)
                                    cleared = []
                                    # J:ES (10-149)
                                    for col in range(10, 150):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # EU:FE (151-161)
                                    for col in range(151, 162):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    # FG:FN (163-170)
                                    for col in range(163, 171):
                                        val = dest_ws_com.Cells(check_row, col).Value
                                        if is_one(val):
                                            dest_ws_com.Cells(check_row, col).Value = None
                                            cleared.append(str(col))
                                    if cleared:
                                        print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус СЗЧ, залишено тільки FF)")
                                
                                else:
                                    # Якщо статус БЗ або інший (не ВЛК/Ш/В/БКБП/КЗВ/ВД/СП/КР/БЧ/РАО/ДИС/РБ/БЗН/СЗЧ)
                                    if stroiovka_status == 'БЗ':
                                        # Для БЗ:
                                        # - Прапорці в J:ES (10-149) — це ОК, залишаємо
                                        # - Прапорці в EU:FJ (151-166) — помилка, скидаємо і ставимо ES=1
                                        
                                        # Перевіряємо чи є прапорці в EU:FJ (151-166)
                                        has_eu_fn_flags = any(is_one(val) for val in flags_eu_fn)
                                        
                                        if has_eu_fn_flags:
                                            # ПОМИЛКА: є прапорці в EU:FJ → скидаємо їх, ставимо ES=1
                                            cleared_eu_fn = []
                                            for col in range(151, 167):  # EU:FJ
                                                val = dest_ws_com.Cells(check_row, col).Value
                                                if is_one(val):
                                                    dest_ws_com.Cells(check_row, col).Value = None
                                                    cleared_eu_fn.append(str(col))
                                            
                                            # Ставимо ES=1
                                            dest_ws_com.Cells(check_row, 149).Value = 1
                                            print(f"   ✓ Рядок {check_row}: ПОМИЛКА БЗ — скинуто {len(cleared_eu_fn)} прапорців з EU:FJ, встановлено ES=1")
                                        else:
                                            # Все ОК: прапорці тільки в J:ES або взагалі немає
                                            # Нічого не робимо, залишаємо як є
                                            pass
                                    else:
                                        # Інші статуси (не ВЛК/Ш/В/БКБП/КЗВ/ВД/СП/КР/БЧ/РАО/ДИС/РБ/БЗН/СЗЧ/БЗ) - скидаємо всі прапорці J:ES (10-149) + EU:FN (151-170)
                                        cleared = []
                                        # J:ES (10-149)
                                        for col in range(10, 150):
                                            val = dest_ws_com.Cells(check_row, col).Value
                                            if is_one(val):
                                                dest_ws_com.Cells(check_row, col).Value = None
                                                cleared.append(str(col))
                                        # EU:FN (151-170)
                                        for col in range(151, 171):
                                            val = dest_ws_com.Cells(check_row, col).Value
                                            if is_one(val):
                                                dest_ws_com.Cells(check_row, col).Value = None
                                                cleared.append(str(col))
                                    if cleared:
                                        print(f"   ✓ Рядок {check_row}: скинуто {len(cleared)} прапорців (статус «{stroiovka_status}» не потребує прапорців)")
                        
                        except Exception as flag_error:
                            # Якщо помилка при проставлянні прапорців, логуємо та продовжуємо
                            print(f"   ⚠️ Помилка при проставлянні прапорців для рядка {check_row}: {flag_error}")

                        # Додаткове правило: якщо EU = 1, статус повинен бути «ВЛК»
                        def ensure_status(expected: str, flag_label: str):
                            if pib_str in stroiovka_data:
                                stroiovka_status = stroiovka_data[pib_str]
                                if stroiovka_status != expected:
                                    stroiovka_errors.append(
                                        f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}», "
                                        f"Статус у стройовці: «{stroiovka_status}» (очікується «{expected}», {flag_label})"
                                    )
                            else:
                                stroiovka_errors.append(
                                    f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}» — не знайдено у стройовці "
                                    f"(очікується статус «{expected}», {flag_label})"
                                )

                        if is_one(eu_val):
                            ensure_status('ВЛК', 'оскільки EU=1')

                        if is_one(ew_val):
                            ensure_status('Ш', 'оскільки EW=1')
                        
                        if is_one(ev_val):
                            # EV більше не використовується для Ш, тільки EW
                            stroiovka_errors.append(
                                f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}», "
                                f"EV=1 (застаріло: для статусу Ш використовуйте тільки EW=1)"
                            )

                        # Для EX/EY перевіряємо не тільки статус В, а й відповідність колонці G
                        if is_one(ex_val):
                            # EX=1 → статус В + G=МВ
                            g_str = str(g_val).strip().upper() if g_val else ''
                            if pib_str in stroiovka_data:
                                stroiovka_status = stroiovka_data[pib_str]
                                # Перевіряємо чи статус починається з "В" (може бути "В" або "В ('МВ)" тощо)
                                if stroiovka_status != 'В' and not stroiovka_status.startswith("В ("):
                                    stroiovka_errors.append(
                                        f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}», "
                                        f"Статус у стройовці: «{stroiovka_status}» (очікується «В», оскільки EX=1)"
                                    )
                                elif g_str != 'МВ':
                                    stroiovka_errors.append(
                                        f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}», "
                                        f"G=«{g_str}» (очікується «МВ» для EX=1, інакше використовуйте EY)"
                                    )
                            else:
                                stroiovka_errors.append(
                                    f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}» — не знайдено у стройовці "
                                    f"(очікується статус «В», оскільки EX=1)"
                                )

                        if is_one(ey_val):
                            # EY=1 → статус В + G≠МВ
                            g_str = str(g_val).strip().upper() if g_val else ''
                            if pib_str in stroiovka_data:
                                stroiovka_status = stroiovka_data[pib_str]
                                # Перевіряємо чи статус починається з "В" (може бути "В" або "В ('МВ)" тощо)
                                if stroiovka_status != 'В' and not stroiovka_status.startswith("В ("):
                                    stroiovka_errors.append(
                                        f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}», "
                                        f"Статус у стройовці: «{stroiovka_status}» (очікується «В», оскільки EY=1)"
                                    )
                                elif g_str == 'МВ':
                                    stroiovka_errors.append(
                                        f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}», "
                                        f"G=«МВ» (для МВ використовуйте EX=1 замість EY)"
                                    )
                            else:
                                stroiovka_errors.append(
                                    f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}» — не знайдено у стройовці "
                                    f"(очікується статус «В», оскільки EY=1)"
                                )

                        if is_one(fc_val):
                            # FC=1 → статус ВД
                            ensure_status('ВД', 'оскільки FC=1')

                        if is_one(fe_val):
                            # FE=1 → статус БЗН
                            ensure_status('БЗН', 'оскільки FE=1')

                        if is_one(ff_val):
                            # FF=1 → статус СЗЧ
                            ensure_status('СЗЧ', 'оскільки FF=1')

                        if is_one(fj_val):
                            # FJ=1 → статус РБ
                            ensure_status('РБ', 'оскільки FJ=1')

                        if is_one(fn_val):
                            # FN=1 → статуси КЗВ/СП/КР/БЧ/РАО/ДИС
                            if pib_str in stroiovka_data:
                                stroiovka_status = stroiovka_data[pib_str]
                                if stroiovka_status not in ['КЗВ', 'СП', 'КР', 'БЧ', 'РАО', 'ДИС']:
                                    stroiovka_errors.append(
                                        f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}», "
                                        f"Статус у стройовці: «{stroiovka_status}» (очікується один з КЗВ/СП/КР/БЧ/РАО/ДИС, оскільки FN=1)"
                                    )
                            else:
                                stroiovka_errors.append(
                                    f"Рядок {check_row} — Підрозділ: «{unit_name}», ПІБ: «{pib_str}» — не знайдено у стройовці "
                                    f"(очікується статус КЗВ/СП/КР/БЧ/РАО/ДИС, оскільки FN=1)"
                                )

                    # Закриваємо Excel (зберігаємо зміни у результуючому файлі)
                    try:
                        stroiovka_wb_com.Close(SaveChanges=False)
                    except Exception as e:
                        print(f"   ⚠️ Помилка при закритті файлу стройовки: {e}")
                    
                    try:
                        # ВАЖЛИВО: Явно зберігаємо файл перед закриттям
                        dest_wb_com.Save()
                        print(f"   💾 Файл збережено з автоматично проставленими прапорцями")
                        dest_wb_com.Close(SaveChanges=False)  # Вже збережено вище
                    except Exception as e:
                        print(f"   ⚠️ Помилка при збереженні/закритті результуючого файлу: {e}")
                        # Спробуємо закрити без збереження
                        try:
                            dest_wb_com.Close(SaveChanges=False)
                        except:
                            pass
                    
                    try:
                        excel.Quit()
                    except Exception as e:
                        print(f"   ⚠️ Помилка при закритті Excel: {e}")

                    print(f"   Перевірено рядків: {checked_count}, знайдено помилок: {len(stroiovka_errors)}")

                    # Відкриваємо файл знову через openpyxl для додавання логів
                    dest_wb = openpyxl.load_workbook(destination_file)
                    dest_ws = dest_wb[target_sheet_name]
                    log_ws = dest_wb['LOG']

                    # Знаходимо останній заповнений рядок у LOG
                    row = log_ws.max_row

                except ImportError:
                    print(f"⚠️ win32com не встановлено. Перевірка стройовки пропущена.")
                    print(f"   Встановіть: pip install pywin32")
                    # Відкриваємо файл знову
                    dest_wb = openpyxl.load_workbook(destination_file)
                    dest_ws = dest_wb[target_sheet_name]
                    log_ws = dest_wb['LOG']
                    row = log_ws.max_row
                except Exception as e:
                    print(f"⚠️ Помилка перевірки стройовки через COM: {e}")
                    import traceback
                    traceback.print_exc()
                    # Відкриваємо файл знову
                    dest_wb = openpyxl.load_workbook(destination_file)
                    dest_ws = dest_wb[target_sheet_name]
                    log_ws = dest_wb['LOG']
                    row = log_ws.max_row

        # Додаємо секцію СТРОЙОВКА до LOG (тільки якщо перевірка була увімкнена)
        if stroiovka_env == '1' and stroiovka_file_path:
            if stroiovka_errors:
                row += 1
                log_ws.cell(row, 1, '')
                row += 1
                log_ws.cell(row, 1, '⚠️ НЕВІДПОВІДНІСТЬ ЗІ СТРОЙОВКОЮ')
                log_ws.cell(row, 1).font = Font(bold=True, size=12, color='FFFFFF')
                log_ws.cell(row, 1).fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
                log_ws.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')
                log_ws.row_dimensions[row].height = 25
                stroiovka_header_row = row
                row += 1

                log_ws.cell(row, 1, f'Знайдено невідповідностей: {len(stroiovka_errors)}')
                log_ws.cell(row, 1).font = Font(bold=True, color='FF8C00')
                row += 1

                # Додаємо перші 100 помилок
                for i, err in enumerate(stroiovka_errors[:100], 1):
                    cell = log_ws.cell(row, 1, f"{i}. {err}")
                    cell.font = Font(size=9, color='C00000')
                    cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    row += 1

                if len(stroiovka_errors) > 100:
                    cell = log_ws.cell(row, 1, f"... і ще {len(stroiovka_errors) - 100} невідповідностей")
                    cell.font = Font(italic=True, color='7F7F7F')
                    row += 1

                log_ws.merge_cells(f'A{stroiovka_header_row}:B{stroiovka_header_row}')
                print(f"⚠️ СТРОЙОВКА: Знайдено невідповідностей: {len(stroiovka_errors)}")
            else:
                row += 1
                log_ws.cell(row, 1, '')
                row += 1
                log_ws.cell(row, 1, '✅ ПЕРЕВІРКА СТРОЙОВКИ: Всі записи відповідають')
                log_ws.cell(row, 1).font = Font(bold=True, color='228B22')
                log_ws.cell(row, 1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                log_ws.merge_cells(f'A{row}:B{row}')
                print(f"✅ СТРОЙОВКА: Невідповідностей не знайдено")

        # === ВИПРАВЛЕННЯ ЗВАНЬ/ПОСАД ===
        correction_logs = []

        # Перевіряємо чи потрібно виправляти звання або посади
        fix_rank = os.environ.get('PY_FIX_RANK', '0') == '1'
        fix_position = os.environ.get('PY_FIX_POSITION', '0') == '1'
        corrections_file = os.environ.get('PY_CORRECTIONS_FILE', '').strip().strip('"').strip("'")

        if (fix_rank or fix_position) and corrections_file and os.path.exists(corrections_file):
            print(f"🔧 Застосування виправлень з файлу: {corrections_file}")

            # Діапазон рядків для виправлення (від початку записаних даних)
            data_start_row = upr_end + 1
            data_end_row = upr_end + written

            if fix_rank:
                print(f"🔧 Виправлення звань (колонка D)...")

                # Колонка C (3) у файлі виправлень = звання
                rank_index = load_corrections_index(corrections_file, value_col=3)

                if rank_index:
                    # Колонка D (4) у результаті = звання
                    rank_logs = apply_corrections(dest_ws, data_start_row, data_end_row,
                                                 rank_index, target_col=4, col_name="звання")
                    correction_logs.extend(rank_logs)

            if fix_position:
                print(f"🔧 Виправлення посад (колонка C)...")

                # Колонка F (6) у файлі виправлень = посада
                position_index = load_corrections_index(corrections_file, value_col=6)

                if position_index:
                    # Колонка C (3) у результаті = посада
                    position_logs = apply_corrections(dest_ws, data_start_row, data_end_row,
                                                     position_index, target_col=3, col_name="посади")
                    correction_logs.extend(position_logs)

            # Додаємо логи виправлень до LOG
            if correction_logs:
                row = log_ws.max_row + 1
                log_ws.cell(row, 1, '')
                row += 1

                # Заголовок секції
                log_ws.cell(row, 1, '🔧 ВИПРАВЛЕННЯ ЗВАНЬ/ПОСАД')
                log_ws.cell(row, 1).font = Font(bold=True, size=12, color='FFFFFF')
                log_ws.cell(row, 1).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                log_ws.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')
                log_ws.row_dimensions[row].height = 25
                log_ws.merge_cells(f'A{row}:B{row}')
                row += 1

                log_ws.cell(row, 1, f'Файл виправлень: {os.path.basename(corrections_file)}')
                log_ws.cell(row, 1).font = Font(size=10, italic=True, color='7F7F7F')
                log_ws.merge_cells(f'A{row}:B{row}')
                row += 1

                # Додаємо всі логи виправлень
                for log_line in correction_logs:
                    cell = log_ws.cell(row, 1, log_line)

                    # Стилі для логів
                    if log_line.strip().startswith('✅'):
                        cell.font = Font(bold=True, color='228B22')
                        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    elif log_line.strip().startswith('ℹ️'):
                        cell.font = Font(color='0070C0')
                        cell.fill = PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid')
                    elif log_line.strip().startswith('✍️'):
                        cell.font = Font(color='4472C4')
                    else:
                        cell.font = Font(color='404040')

                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    log_ws.merge_cells(f'A{row}:B{row}')
                    row += 1

        # 11. Оновлення статусу FO (якщо увімкнено)
        update_status_env = os.environ.get('PY_UPDATE_STATUS', '0')
        if update_status_env == '1' and update_fo_status is not None:
            print(f"\n🔄 Оновлення статусу FO...")
            try:
                # Викликаємо функцію оновлення статусу (сама записує в LOG)
                row += 1  # Пропуск рядка перед секцією
                fo_stats = update_fo_status(dest_ws, log_ws, row)
                row = fo_stats.get('log_end_row', row)
                
                print(f"✅ Статус FO оновлено:")
                print(f"   Оновлено: {fo_stats['updates']}")
                print(f"   Помилок: {fo_stats['errors']}")
                print(f"   Пропущено: {fo_stats['skipped']}")
                    
            except Exception as e:
                print(f"⚠️ Помилка оновлення статусу FO: {e}")
                import traceback
                traceback.print_exc()
                
                # Додаємо помилку до LOG
                if log_ws:
                    row += 1
                    log_ws.cell(row, 1, f'❌ Помилка оновлення статусу FO: {e}')
                    log_ws.cell(row, 1).font = Font(bold=True, color='C00000')
                    log_ws.cell(row, 1).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    log_ws.merge_cells(f'A{row}:B{row}')
                    row += 1

        # 12. Збереження файлу
        print(f"💾 Збереження файлу...")
        dest_wb.save(destination_file)
        dest_wb.close()
        print(f"✅ Готово! Записано рядків: {written}")

        # 13. Автовідкриття (якщо потрібно)
        if auto_open:
            try:
                import subprocess
                if sys.platform == 'win32':
                    os.startfile(destination_file)
                elif sys.platform == 'darwin':
                    subprocess.run(['open', destination_file])
                else:
                    subprocess.run(['xdg-open', destination_file])
            except Exception as e:
                print(f"⚠️ Не вдалося відкрити файл: {e}")

        # 14. Повертаємо результат
        return {
            "ok": True,
            "stats": {
                "filesProcessed": len(files),
                "rowsWritten": written,
                "unitsFound": len(buckets),
                "capacity": capacity,
                "fnpErrors": len(fnp_errors),
                "duplicatesErrors": len(duplicates_errors),
                "stroiovkaErrors": len(stroiovka_errors)
            }
        }
    
    except PermissionError as e:
        error_msg = f"❌ ФАЙЛ ВІДКРИТО В EXCEL!\n\nЗакрийте файл '{destination_file}' в Excel та спробуйте знову.\n\nТехнічна помилка: {str(e)}"
        print(error_msg, file=sys.stderr)
        return {"ok": False, "error": error_msg}
    except Exception as e:
        import traceback
        error_msg = f"Помилка: {str(e)}\n{traceback.format_exc()}"
        print(error_msg, file=sys.stderr)
        return {"ok": False, "error": str(e)}


# ============== CLI ==============

if __name__ == "__main__":
    import argparse
    
    # Встановлюємо UTF-8 для stdout в Windows
    if sys.platform == 'win32':
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, errors='replace')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, errors='replace')
    
    print("🐍 Python скрипт process_dodatok10.py запущено")
    print(f"🐍 Python версія: {sys.version}")
    
    parser = argparse.ArgumentParser(description='Обробка Додатку 10')
    parser.add_argument('--input-folder', required=True, help='Папка з файлами (шукає "Дод.10", "Дод 10", "Додаток 10")')
    parser.add_argument('--destination-file', required=True, help='Цільовий Excel файл')
    parser.add_argument('--auto-open', action='store_true', help='Відкрити файл після обробки')
    parser.add_argument('--no-formula-cols', action='store_true', help='Перезаписати формульні колонки')
    
    args = parser.parse_args()
    
    print(f"📂 Вхідна папка: {args.input_folder}")
    print(f"💾 Файл призначення: {args.destination_file}")
    print(f"🚀 Автовідкриття: {args.auto_open}")
    print(f"🔧 Ігнорувати формульні колонки: {not args.no_formula_cols}")
    print()
    
    result = process_dodatok10(
        input_folder=args.input_folder,
        destination_file=args.destination_file,
        auto_open=args.auto_open,
        ignore_formula_cols=not args.no_formula_cols
    )
    
    if result["ok"]:
        print("\n✅ УСПІШНО")
        # ВАЖЛИВО: Формат для парсингу в electron/main.ts
        print(f"Знайдено файлів: {result['stats']['filesProcessed']}")
        print(f"Записано рядків: {result['stats']['rowsWritten']}")
        print(f"Підрозділів знайдено: {result['stats']['unitsFound']}")
        sys.exit(0)
    else:
        print(f"\n❌ ПОМИЛКА: {result['error']}", file=sys.stderr)
        sys.exit(1)
