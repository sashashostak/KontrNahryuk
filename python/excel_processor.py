"""
Excel Processor для програми Стройовка
Використовує openpyxl для безпечної роботи з Excel файлами
"""

import sys
import json
import io
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill
from excel_sanitizer import sanitize_cells
from excel_mismatches import check_mismatches
from excel_slice_check import check_slices
from excel_duplicates import check_duplicates
from text_utils import normalize_text, normalize_list

# Встановлюємо UTF-8 для stdout/stderr на Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


class ProcessingLog:
    """Клас для збору логів обробки"""
    
    def __init__(self):
        self.entries = []
        self.start_time = datetime.now()
        
    def add_entry(self, operation: str, source_file: str, sheet: str, rows: int, details: str = ""):
        """Додати запис до логу"""
        self.entries.append({
            'operation': operation,
            'source_file': source_file,
            'sheet': sheet,
            'rows': rows,
            'details': details,
            'is_separator': False
        })
    
    def add_separator(self, title: str):
        """Додати розділювач етапів"""
        self.entries.append({
            'operation': f'═══ {title} ═══',
            'source_file': '',
            'sheet': '',
            'rows': '',
            'details': '',
            'is_separator': True
        })
    
    def get_summary(self) -> Dict:
        """Отримати підсумок обробки"""
        # Фільтруємо розділювачі та підсумовуємо тільки числові значення
        total_rows = sum(
            entry['rows'] for entry in self.entries 
            if not entry.get('is_separator', False) and isinstance(entry['rows'], (int, float))
        )
        files_processed = len(set(entry['source_file'] for entry in self.entries if entry['source_file']))
        
        # Рахуємо тільки звичайні операції (без розділювачів)
        operations = sum(1 for entry in self.entries if not entry.get('is_separator', False))
        
        return {
            'total_rows': total_rows,
            'files_processed': files_processed,
            'operations': operations,
            'duration': (datetime.now() - self.start_time).total_seconds()
        }


class ExcelProcessor:
    """Процесор для копіювання даних між Excel файлами"""
    
    def __init__(self):
        self.destination_wb = None
        self.destination_path = None
        self.log = ProcessingLog()
        
    def load_destination(self, file_path: str) -> None:
        """Завантажити файл призначення"""
        print(f"📖 Завантаження файлу призначення: {file_path}")
        
        # Перевіряємо чи файл не заблокований перед завантаженням
        try:
            # Спроба відкрити файл для запису (перевірка блокування)
            with open(file_path, 'r+b') as f:
                pass
        except PermissionError:
            error_msg = (
                f"❌ Файл '{file_path}' відкритий в іншій програмі!\n\n"
                f"🔧 Закрийте Excel та спробуйте ще раз."
            )
            print(error_msg, file=sys.stderr)
            raise PermissionError(error_msg)
        
        self.destination_path = file_path
        self.destination_wb = load_workbook(file_path)
        print(f"✅ Файл завантажено, листів: {len(self.destination_wb.sheetnames)}")
        
    def build_index(self, sheet_name: str, key_column: str, blacklist: List[str]) -> Dict[str, List[int]]:
        """
        Побудувати індекс підрозділів у файлі призначення

        Returns:
            Dict з ключами підрозділів та списками рядків де вони зустрічаються
        """
        print(f"\n🔍 Сканування листа '{sheet_name}'...")

        if sheet_name not in self.destination_wb.sheetnames:
            print(f"⚠️ Лист '{sheet_name}' не знайдено")
            return {}

        sheet = self.destination_wb[sheet_name]
        index = {}

        # Конвертуємо літеру колонки в номер (B=2, C=3)
        col_num = ord(key_column.upper()) - ord('A') + 1

        print(f"   Колонка ключа: {key_column} (номер {col_num})")

        # Нормалізуємо blacklist один раз
        normalized_blacklist = normalize_list(blacklist, remove_spaces=True)

        found_count = 0
        for row_num in range(2, sheet.max_row + 1):  # Пропускаємо заголовок
            cell = sheet.cell(row=row_num, column=col_num)
            value = cell.value

            if value:
                # Нормалізація ключа через єдину функцію
                key = normalize_text(value, remove_spaces=True)

                # Перевірка blacklist (вже нормалізований)
                if key in normalized_blacklist:
                    continue

                if key not in index:
                    index[key] = []
                    found_count += 1

                index[key].append(row_num)

        print(f"   📊 Знайдено {found_count} унікальних підрозділів")
        return index
        
    def find_contiguous_block(self, sheet: Worksheet, key_column: str,
                             key: str, start_row: int = 2) -> Optional[Tuple[int, int]]:
        """
        Знайти контігуальний блок рядків з заданим ключем

        Returns:
            Tuple (start_row, end_row) або None
        """
        col_num = ord(key_column.upper()) - ord('A') + 1

        # Шукаємо перший рядок
        first_row = None
        for row_num in range(start_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=col_num).value
            if cell_value:
                # Використовуємо єдину функцію нормалізації
                normalized = normalize_text(cell_value, remove_spaces=True)
                if normalized == key:
                    first_row = row_num
                    break

        if first_row is None:
            return None

        # Знаходимо останній рядок блоку
        last_row = first_row
        for row_num in range(first_row + 1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=col_num).value
            if cell_value:
                # Використовуємо єдину функцію нормалізації
                normalized = normalize_text(cell_value, remove_spaces=True)
                if normalized == key:
                    last_row = row_num
                else:
                    break
            else:
                # Порожня комірка - кінець блоку
                break

        return (first_row, last_row)
        
    def copy_data(self, source_file: str, dest_sheet_name: str, 
                  key_column: str, data_columns: List[str], index: Dict[str, List[int]]) -> int:
        """
        Копіювати дані з вхідного файлу
        
        Returns:
            Кількість скопійованих рядків
        """
        print(f"\n📄 Обробка файлу: {Path(source_file).name}")
        
        source_wb = load_workbook(source_file, data_only=True)
        
        # Логування списку листів у файлі (особливо для 1241)
        file_name = Path(source_file).name
        if '1241' in file_name:
            print(f"   📋 Листи у файлі 1241: {source_wb.sheetnames}")
        
        # Шукаємо лист з урахуванням можливих пробілів в назві
        actual_sheet_name = None
        for sheet_name in source_wb.sheetnames:
            if sheet_name.strip() == dest_sheet_name.strip():
                actual_sheet_name = sheet_name
                break
        
        if actual_sheet_name is None:
            print(f"   ⚠️ Лист '{dest_sheet_name}' не знайдено у вхідному файлі")
            print(f"   📋 Доступні листи: {source_wb.sheetnames}")
            return 0
        
        # Якщо назва листа відрізняється (пробіли), повідомляємо про це
        if actual_sheet_name != dest_sheet_name:
            print(f"   🔧 Знайдено лист '{actual_sheet_name}' (з пробілами) замість '{dest_sheet_name}'")
        
        source_sheet = source_wb[actual_sheet_name]
        dest_sheet = self.destination_wb[dest_sheet_name]
        
        # Конвертуємо літери колонок в номери
        key_col_num = ord(key_column.upper()) - ord('A') + 1
        data_col_nums = [ord(col.upper()) - ord('A') + 1 for col in data_columns]
        
        copied_rows = 0
        
        # Проходимо по всіх ключах з індексу
        for key in index.keys():
            # Знаходимо блок у джерелі
            source_block = self.find_contiguous_block(source_sheet, key_column, key)
            
            if not source_block:
                continue
            
            src_start, src_end = source_block
            src_count = src_end - src_start + 1
            
            print(f"\n   --- Підрозділ: '{key}' ---")
            print(f"   ✅ Джерело: рядки {src_start}-{src_end} ({src_count} рядків)")
            
            # Знаходимо блоки у призначенні
            dest_rows = index[key]
            if not dest_rows:
                print(f"   ⚠️ Не знайдено у призначенні")
                continue
            
            # Групуємо у контігуальні блоки
            dest_blocks = self._group_contiguous(dest_rows)
            print(f"   ✅ Призначення: {len(dest_blocks)} блок(ів)")
            
            # Копіюємо дані
            src_row_ptr = src_start
            
            for block_idx, dest_block in enumerate(dest_blocks):
                dest_start = dest_block[0]
                dest_end = dest_block[-1]
                dest_count = len(dest_block)
                
                # Скільки рядків можемо скопіювати
                rows_left = src_end - src_row_ptr + 1
                if rows_left <= 0:
                    break
                    
                rows_to_copy = min(rows_left, dest_count)
                
                print(f"      Блок {block_idx + 1}: копіювання {rows_to_copy} рядків")
                
                # Копіюємо по рядкам
                for i in range(rows_to_copy):
                    src_row = src_row_ptr + i
                    dest_row = dest_start + i
                    
                    # Копіюємо кожну колонку
                    for col_num in data_col_nums:
                        value = source_sheet.cell(row=src_row, column=col_num).value
                        dest_sheet.cell(row=dest_row, column=col_num).value = value
                    
                    copied_rows += 1
                
                # Очищаємо хвіст якщо є
                if rows_to_copy < dest_count:
                    for i in range(rows_to_copy, dest_count):
                        dest_row = dest_start + i
                        for col_num in data_col_nums:
                            dest_sheet.cell(row=dest_row, column=col_num).value = None
                
                src_row_ptr += rows_to_copy
        
        print(f"   ✅ Скопійовано {copied_rows} рядків")
        
        # Додаємо в лог
        if copied_rows > 0:
            self.log.add_entry(
                operation='Копіювання даних',
                source_file=Path(source_file).name,
                sheet=dest_sheet_name,
                rows=copied_rows,
                details=f"Ключ: {key_column}, Колонки: {', '.join(data_columns)}"
            )
        
        return copied_rows
        
    def _group_contiguous(self, rows: List[int]) -> List[List[int]]:
        """Групувати рядки у контігуальні блоки"""
        if not rows:
            return []
        
        sorted_rows = sorted(rows)
        blocks = []
        current_block = [sorted_rows[0]]
        
        for row in sorted_rows[1:]:
            if row == current_block[-1] + 1:
                current_block.append(row)
            else:
                blocks.append(current_block)
                current_block = [row]
        
        blocks.append(current_block)
        return blocks
        
    def create_log_sheet(self) -> None:
        """Створити лист LOG з інформацією про обробку"""
        print(f"\n📋 Створення листа LOG...")
        
        # Видаляємо старий лист LOG якщо існує
        if 'LOG' in self.destination_wb.sheetnames:
            del self.destination_wb['LOG']
        
        # Створюємо новий лист
        log_sheet = self.destination_wb.create_sheet('LOG', 0)  # Перша позиція
        
        # Стилі
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14)
        
        # Заголовок
        log_sheet['A1'] = 'ЛОГ ОБРОБКИ ФАЙЛІВ'
        log_sheet['A1'].font = title_font
        log_sheet.merge_cells('A1:F1')
        
        # Загальна інформація
        summary = self.log.get_summary()
        row = 3
        log_sheet[f'A{row}'] = 'Дата обробки:'
        log_sheet[f'B{row}'] = self.log.start_time.strftime('%d.%m.%Y %H:%M:%S')
        log_sheet[f'B{row}'].font = Font(bold=True)
        
        row += 1
        log_sheet[f'A{row}'] = 'Всього файлів оброблено:'
        log_sheet[f'B{row}'] = summary['files_processed']
        log_sheet[f'B{row}'].font = Font(bold=True)
        
        row += 1
        log_sheet[f'A{row}'] = 'Всього операцій:'
        log_sheet[f'B{row}'] = summary['operations']
        log_sheet[f'B{row}'].font = Font(bold=True)
        
        row += 1
        log_sheet[f'A{row}'] = 'Всього рядків скопійовано:'
        log_sheet[f'B{row}'] = summary['total_rows']
        log_sheet[f'B{row}'].font = Font(bold=True, color="00B050")
        
        row += 1
        log_sheet[f'A{row}'] = 'Час обробки (сек):'
        log_sheet[f'B{row}'] = f"{summary['duration']:.2f}"
        log_sheet[f'B{row}'].font = Font(bold=True)
        
        # Таблиця операцій
        row += 2
        headers = ['№', 'Операція', 'Файл джерело', 'Лист', 'Рядків', 'Деталі']
        for col_idx, header in enumerate(headers, start=1):
            cell = log_sheet.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Дані
        entry_number = 0  # Лічильник для звичайних записів (без розділювачів)
        for entry in self.log.entries:
            row += 1
            
            # Перевірка чи це розділювач
            if entry.get('is_separator', False):
                # Стиль для розділювача
                cell = log_sheet.cell(row=row, column=2)
                cell.value = entry['operation']
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                log_sheet.merge_cells(f'B{row}:F{row}')
                
                # Порожня клітинка №
                log_sheet.cell(row=row, column=1).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            else:
                # Звичайний запис
                entry_number += 1
                log_sheet.cell(row=row, column=1).value = entry_number
                log_sheet.cell(row=row, column=2).value = entry['operation']
                log_sheet.cell(row=row, column=3).value = entry['source_file']
                log_sheet.cell(row=row, column=4).value = entry['sheet']
                log_sheet.cell(row=row, column=5).value = entry['rows']
                log_sheet.cell(row=row, column=6).value = entry['details']
        
        # Автоматичне підлаштування ширини колонок
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            max_length = 0
            
            for cell in log_sheet[col_letter]:
                try:
                    # Пропускаємо MergedCell та порожні клітинки
                    if hasattr(cell, 'value') and cell.value:
                        # Враховуємо довжину тексту
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Встановлюємо ширину з невеликим запасом
            adjusted_width = min(max_length + 2, 100)  # Максимум 100 для дуже довгих рядків
            
            # Мінімальна ширина для кожної колонки
            min_widths = {
                'A': 5,   # №
                'B': 20,  # Операція
                'C': 15,  # Файл джерело
                'D': 15,  # Лист
                'E': 10,  # Рядків
                'F': 30   # Деталі
            }
            
            final_width = max(adjusted_width, min_widths.get(col_letter, 10))
            log_sheet.column_dimensions[col_letter].width = final_width
        
        print(f"✅ Лист LOG створено з {len(self.log.entries)} записами")
    
    def sanitize_sheets(self, sheet_names: List[str]) -> Dict[str, int]:
        """
        Санітизувати дані у вказаних листах (колонки A:H)
        
        Args:
            sheet_names: Список назв листів для санітизації
        
        Returns:
            Словник {sheet_name: кількість_змін}
        """
        print(f"\n🧹 === САНІТИЗАЦІЯ ДАНИХ (КОЛОНКИ A:H) ===")
        print(f"   Листи: {', '.join(sheet_names)}")
        
        results = {}
        
        for sheet_name in sheet_names:
            # Шукаємо лист з урахуванням пробілів
            actual_sheet_name = None
            for name in self.destination_wb.sheetnames:
                if name.strip() == sheet_name.strip():
                    actual_sheet_name = name
                    break
            
            if actual_sheet_name is None:
                print(f"   ⚠️ Лист '{sheet_name}' не знайдено")
                continue
            
            sheet = self.destination_wb[actual_sheet_name]
            print(f"\n   📄 Обробка листа: {actual_sheet_name}")
            
            # Санітизуємо тільки колонки A:H
            cells = []
            for row in sheet.iter_rows(min_col=1, max_col=8):  # A=1, H=8
                for cell in row:
                    cells.append(cell)
            
            total, changed, preview = sanitize_cells(cells, show_preview=True)
            results[actual_sheet_name] = changed
            
            print(f"      Всього клітинок: {total}")
            print(f"      Змінено: {changed}")
            
            if preview:
                print(f"      Превʼю змін (перші {len(preview)}):")
                for change in preview[:10]:  # Показуємо перші 10
                    print(f"         {change['cell']}: {change['old']} → {change['new']}")
            
            # Додаємо в лог
            if changed > 0:
                self.log.add_entry(
                    operation='Санітизація даних (A:H)',
                    source_file='',
                    sheet=actual_sheet_name,
                    rows=changed,
                    details=f'Очищено {changed} клітинок з {total} (колонки A:H)'
                )
        
        total_changes = sum(results.values())
        print(f"\n   ✅ Всього змін: {total_changes}")
        
        return results
    
    def check_mismatches(self) -> Tuple[List, Dict]:
        """
        Перевірити невідповідності між листами ЗС та БЗ
        
        Returns:
            (список_невідповідностей, статистика)
        """
        print(f"\n🔍 === ПЕРЕВІРКА НЕВІДПОВІДНОСТЕЙ ===")
        
        mismatches, error, stats = check_mismatches(self.destination_path)
        
        if error:
            print(f"   ❌ Помилка: {error}")
            return [], {}
        
        # Додаємо розділювач в лог
        self.log.add_separator('ПЕРЕВІРКА НЕВІДПОВІДНОСТЕЙ')
        
        # Додаємо підсумкову інформацію
        if stats.get('total', 0) > 0:
            self.log.add_entry(
                operation='Знайдено невідповідностей',
                source_file='',
                sheet='ЗС / БЗ',
                rows=stats['total'],
                details=f"ЗС→БЗ: {stats['s1_missing_in_s2']}, БЗ→ЗС: {stats['s2_missing_in_s1']}"
            )
            
            # Додаємо детальну інформацію про кожну невідповідність
            for mismatch in mismatches:
                self.log.add_entry(
                    operation='Невідповідність',
                    source_file='',
                    sheet=mismatch.sheet,
                    rows='',
                    details=f"{mismatch.cell_addr}: «{mismatch.value}» — {mismatch.reason}"
                )
        else:
            self.log.add_entry(
                operation='Невідповідностей не знайдено',
                source_file='',
                sheet='ЗС / БЗ',
                rows=0,
                details='Всі дані збігаються'
            )
        
        return mismatches, stats
    
    def check_slices(self) -> Tuple[List, Dict]:
        """
        Перевірити "зрізи" (колонки F/G)
        
        Returns:
            (список_проблем, статистика)
        """
        print(f"\n🔍 === ПЕРЕВІРКА «ЗРІЗІВ» (F/G) ===")
        
        issues, error, stats = check_slices(self.destination_path)
        
        if error:
            print(f"   ❌ Помилка: {error}")
            return [], {}
        
        # Додаємо розділювач в лог
        self.log.add_separator('ПЕРЕВІРКА «ЗРІЗІВ» (F/G)')
        
        # Додаємо підсумкову інформацію
        if stats.get('total', 0) > 0:
            self.log.add_entry(
                operation='Знайдено проблем з «зрізами»',
                source_file='ЗС',
                sheet='',
                rows=stats['total'],
                details=f"У F є токен (Ш/ВЛК/ВД), але G порожня"
            )
            
            # Додаємо детальну інформацію про кожну проблему
            for issue in issues:
                self.log.add_entry(
                    operation='Проблема з «зрізом»',
                    source_file=issue.subunit,   # Колонка C LOG = Колонка B ЗС (підрозділ)
                    sheet=issue.fio,             # Колонка D LOG = Колонка D ЗС (ПІБ)
                    rows='',
                    details=f"Псевдо: {issue.pseudo} | Рядок {issue.row} | F={issue.val_f} | G=порожньо"
                )
        else:
            self.log.add_entry(
                operation='Проблем не знайдено',
                source_file='ЗС',
                sheet='',
                rows=0,
                details='Всі «зрізи» заповнені правильно'
            )
        
        return issues, stats
    
    def check_duplicates(self) -> Tuple[List, Dict]:
        """
        Перевірити дублікати ПІБ в листах
        
        Returns:
            (список_дублікатів, статистика)
        """
        print(f"\n🔍 === ПЕРЕВІРКА ДУБЛІКАТІВ ===")
        
        duplicates, error, stats = check_duplicates(self.destination_path)
        
        if error:
            print(f"   ❌ Помилка: {error}")
            return [], {}
        
        # Додаємо розділювач в лог
        self.log.add_separator('ПЕРЕВІРКА ДУБЛІКАТІВ')
        
        # Додаємо підсумкову інформацію
        if stats.get('total', 0) > 0:
            self.log.add_entry(
                operation='Знайдено дублікатів',
                source_file='',
                sheet='ЗС / БЗ / 3бСпП БЗ',
                rows=stats['total'],
                details=f"Повторюються значення в колонках D (ЗС) та E (БЗ, 3бСпП БЗ)"
            )
            
            # Додаємо детальну інформацію про кожен дублікат
            for dup in duplicates:
                rows_str = ", ".join(map(str, dup.rows))
                self.log.add_entry(
                    operation='Дублікат',
                    source_file=dup.sheet,
                    sheet=f"Колонка {dup.column}",
                    rows=dup.count,
                    details=f"«{dup.value}» — рядки: {rows_str}"
                )
        else:
            self.log.add_entry(
                operation='Дублікатів не знайдено',
                source_file='',
                sheet='ЗС / БЗ / 3бСпП БЗ',
                rows=0,
                details='Всі значення унікальні'
            )
        
        return duplicates, stats
    
    def save(self) -> None:
        """Зберегти файл призначення"""
        print(f"\n💾 Збереження файлу...")
        
        # Створюємо лист LOG перед збереженням
        self.create_log_sheet()
        
        try:
            self.destination_wb.save(self.destination_path)
            print(f"✅ Файл збережено: {self.destination_path}")
        except PermissionError as e:
            error_msg = (
                f"❌ Неможливо зберегти файл!\n\n"
                f"Файл '{self.destination_path}' відкритий в іншій програмі (Excel).\n\n"
                f"🔧 Закрийте файл та спробуйте ще раз."
            )
            print(error_msg, file=sys.stderr)
            raise PermissionError(error_msg) from e

    
    def copy_range_direct(self, source_file: str, source_sheet: str, 
                         dest_sheet: str, cell_range: str) -> int:
        """
        Пряме копіювання діапазону з одного листа в інший
        
        Args:
            source_file: Шлях до вхідного файлу
            source_sheet: Назва листа у вхідному файлі
            dest_sheet: Назва листа у файлі призначення
            cell_range: Діапазон комірок, напр. "C4:H231"
        
        Returns:
            Кількість скопійованих комірок
        """
        print(f"\n📋 Пряме копіювання діапазону {cell_range}")
        print(f"   З: {source_file}")
        print(f"   Лист джерела: {source_sheet}")
        print(f"   Лист призначення: {dest_sheet}")
        
        try:
            # Завантажуємо вхідний файл
            src_wb = load_workbook(source_file, data_only=True)
            
            # Шукаємо лист джерела з урахуванням пробілів
            actual_source_sheet = None
            for sheet_name in src_wb.sheetnames:
                if sheet_name.strip() == source_sheet.strip():
                    actual_source_sheet = sheet_name
                    break
            
            if actual_source_sheet is None:
                print(f"   ⚠️ Лист '{source_sheet}' не знайдено у вхідному файлі")
                print(f"   📋 Доступні листи: {src_wb.sheetnames}")
                return 0
            
            # Шукаємо лист призначення з урахуванням пробілів
            actual_dest_sheet = None
            for sheet_name in self.destination_wb.sheetnames:
                if sheet_name.strip() == dest_sheet.strip():
                    actual_dest_sheet = sheet_name
                    break
            
            if actual_dest_sheet is None:
                print(f"   ⚠️ Лист '{dest_sheet}' не знайдено у файлі призначення")
                print(f"   📋 Доступні листи: {self.destination_wb.sheetnames}")
                return 0
            
            # Повідомляємо про різницю в назвах
            if actual_source_sheet != source_sheet:
                print(f"   🔧 Джерело: знайдено '{actual_source_sheet}' замість '{source_sheet}'")
            if actual_dest_sheet != dest_sheet:
                print(f"   🔧 Призначення: знайдено '{actual_dest_sheet}' замість '{dest_sheet}'")
            
            src_ws = src_wb[actual_source_sheet]
            dest_ws = self.destination_wb[actual_dest_sheet]
            
            # Копіюємо значення з діапазону
            cells_copied = 0
            for row in src_ws[cell_range]:
                for cell in row:
                    dest_cell = dest_ws[cell.coordinate]
                    dest_cell.value = cell.value
                    cells_copied += 1
            
            print(f"   ✅ Скопійовано {cells_copied} комірок")
            
            # Додаємо в лог
            if cells_copied > 0:
                # Розраховуємо кількість рядків з діапазону
                range_parts = cell_range.split(':')
                if len(range_parts) == 2:
                    start_row = int(''.join(filter(str.isdigit, range_parts[0])))
                    end_row = int(''.join(filter(str.isdigit, range_parts[1])))
                    rows_count = end_row - start_row + 1
                else:
                    rows_count = cells_copied
                
                self.log.add_entry(
                    operation='Копіювання діапазону (3БСП)',
                    source_file=Path(source_file).name,
                    sheet=source_sheet,
                    rows=rows_count,
                    details=f"Діапазон: {cell_range} ({cells_copied} комірок)"
                )
            
            return cells_copied
            
        except Exception as e:
            print(f"   ❌ Помилка: {str(e)}")
            return 0


def main():
    """Головна функція - приймає JSON конфігурацію через файл або stdin"""
    
    try:
        # Перевіряємо чи передано шлях до конфігурації як аргумент
        if len(sys.argv) > 1:
            # Читаємо з файлу (надійніший спосіб для Unicode шляхів)
            config_file_path = sys.argv[1]
            print(f"📖 Читання конфігурації з файлу: {config_file_path}\n")
            
            with open(config_file_path, 'r', encoding='utf-8-sig') as f:
                config_json = f.read()
            config = json.loads(config_json)
        else:
            # Читаємо з stdin (старий спосіб, залишено для сумісності)
            config_json = sys.stdin.read()
            
            # Намагаємось виправити кодування якщо потрібно
            try:
                config = json.loads(config_json)
            except (json.JSONDecodeError, UnicodeDecodeError):
                # Якщо не вдалося, пробуємо перекодувати
                config_json = config_json.encode('latin1').decode('utf-8')
                config = json.loads(config_json)
        
        print("🚀 === PYTHON EXCEL PROCESSOR ===\n")
        
        processor = ExcelProcessor()
        
        # Завантажуємо файл призначення
        processor.load_destination(config['destination_file'])
        
        # Обробляємо кожен лист
        total_copied = 0
        
        # === СПЕЦІАЛЬНА ОБРОБКА: 3БСП ===
        print(f"\n🔍 Перевірка enable_3bsp: {config.get('enable_3bsp', False)}")
        if config.get('enable_3bsp', False):
            processor.log.add_separator("СПЕЦІАЛЬНА ОБРОБКА: 3БСП")
            
            print("\n🔧 === СПЕЦІАЛЬНА ОБРОБКА: 3БСП ===")
            print(f"   Параметр enable_3bsp: {config.get('enable_3bsp')}")
            print(f"   Кількість файлів для перевірки: {len(config['source_files'])}")
            
            found_3bsp_files = []
            for source_file in config['source_files']:
                file_name = source_file.split('\\')[-1].split('/')[-1]
                file_name_lower = file_name.lower()
                
                # Логування кожного файлу
                if '3' in file_name and ('б' in file_name_lower or 'b' in file_name_lower):
                    print(f"   🔎 Перевірка файлу: {file_name}")
                    print(f"      Нижній регістр: {file_name_lower}")
                    print(f"      Містить '3бспп': {'3бспп' in file_name_lower}")
                
                # Шукаємо різні варіанти написання: 3БСпП, 3бСпП, 3бспп
                if '3бспп' in file_name_lower:
                    found_3bsp_files.append(file_name)
                    print(f"\n📄 ✅ Знайдено файл 3БСП: {file_name}")
                    cells_copied = processor.copy_range_direct(
                        source_file,
                        source_sheet='3бСпП БЗ',
                        dest_sheet='3бСпП БЗ',
                        cell_range='C4:H231'
                    )
                    total_copied += cells_copied
                    print(f"   ✅ 3БСП: скопійовано {cells_copied} комірок")
            
            print(f"\n   📊 Знайдено файлів 3БСП: {len(found_3bsp_files)}")
            if found_3bsp_files:
                for f in found_3bsp_files:
                    print(f"      - {f}")
            print(f"   📊 Всього скопійовано в режимі 3БСП: {total_copied} комірок")
        else:
            print("   ⚠️ Параметр enable_3bsp = False, пропускаємо обробку 3БСП")
        
        # === ОСНОВНА ОБРОБКА ЛИСТІВ ===
        for sheet_config in config['sheets']:
            sheet_name = sheet_config['name']
            key_column = sheet_config['key_column']
            data_columns = sheet_config['data_columns']
            blacklist = sheet_config.get('blacklist', [])
            
            processor.log.add_separator(f"ОБРОБКА ЛИСТА: {sheet_name}")
            
            print(f"\n📋 === ОБРОБКА ЛИСТА: {sheet_name} ===")
            print(f"   Ключ: {key_column}, Дані: {data_columns}")
            
            # Будуємо індекс
            index = processor.build_index(sheet_name, key_column, blacklist)
            
            if not index:
                print(f"   ⚠️ Індекс порожній, пропускаємо")
                continue
            
            # Обробляємо всі вхідні файли
            for source_file in config['source_files']:
                # === СПЕЦІАЛЬНА ЛОГІКА ДЛЯ ФАЙЛУ 1241 ===
                file_name = source_file.split('\\')[-1].split('/')[-1]
                if '1241' in file_name:
                    # Файл 1241 обробляється ТІЛЬКИ для листа БЗ з колонкою C
                    if sheet_name == 'ЗС':
                        print(f"   🔧 Файл 1241: пропускаємо лист ЗС (обробляється тільки БЗ)")
                        continue
                    elif sheet_name == 'БЗ':
                        print(f"   🔧 Файл 1241: використовую колонку C для БЗ")
                        special_key_column = 'C'
                        copied = processor.copy_data(
                            source_file, 
                            sheet_name, 
                            special_key_column,  # Колонка C для 1241
                            data_columns, 
                            index
                        )
                        total_copied += copied
                        continue
                
                # Стандартна обробка
                copied = processor.copy_data(
                    source_file, 
                    sheet_name, 
                    key_column, 
                    data_columns, 
                    index
                )
                total_copied += copied
        
        # === САНІТИЗАЦІЯ ДАНИХ ===
        if config.get('enable_sanitizer', False):
            processor.log.add_separator("САНІТИЗАЦІЯ ДАНИХ")
            
            print("\n🧹 === САНІТИЗАЦІЯ УВІМКНЕНА ===")
            # Збираємо назви всіх оброблених листів
            sheet_names = [sheet_config['name'] for sheet_config in config['sheets']]
            
            # Додаємо спеціальні листи для 3БСП
            if config.get('enable_3bsp', False):
                sheet_names.append('3бСпП БЗ')
            
            # Виконуємо санітизацію
            sanitize_results = processor.sanitize_sheets(sheet_names)
            print(f"   ✅ Санітизація завершена: {sum(sanitize_results.values())} змін")
        else:
            print("\n   ⚠️ Параметр enable_sanitizer = False, пропускаємо санітизацію")
        
        # Створюємо лист LOG
        processor.create_log_sheet()
        
        # Зберігаємо файл (потрібно для перевірок, які читають з диска)
        processor.save()
        
        # === ПЕРЕВІРКА НЕВІДПОВІДНОСТЕЙ ===
        if config.get('enable_mismatches', False):
            print("\n🔍 === ПЕРЕВІРКА НЕВІДПОВІДНОСТЕЙ УВІМКНЕНА ===")
            mismatches, mismatch_stats = processor.check_mismatches()
            
            if mismatch_stats.get('total', 0) > 0:
                print(f"   ⚠️ Знайдено невідповідностей: {mismatch_stats['total']}")
                print(f"      ЗС → відсутні в БЗ: {mismatch_stats['s1_missing_in_s2']}")
                print(f"      БЗ → відсутні в ЗС: {mismatch_stats['s2_missing_in_s1']}")
            else:
                print(f"   ✅ Невідповідностей не знайдено")
        else:
            print("\n   ⚠️ Параметр enable_mismatches = False, пропускаємо перевірку")
        
        # === ПЕРЕВІРКА «ЗРІЗІВ» (F/G) ===
        if config.get('enable_slice_check', False):
            print("\n🔍 === ПЕРЕВІРКА «ЗРІЗІВ» УВІМКНЕНА ===")
            slice_issues, slice_stats = processor.check_slices()
            
            if slice_stats.get('total', 0) > 0:
                print(f"   ⚠️ Знайдено проблем з «зрізами»: {slice_stats['total']}")
            else:
                print(f"   ✅ Проблем не знайдено")
        else:
            print("\n   ⚠️ Параметр enable_slice_check = False, пропускаємо перевірку")
        
        # === ПЕРЕВІРКА ДУБЛІКАТІВ ===
        if config.get('enable_duplicates', False):
            print("\n🔍 === ПЕРЕВІРКА ДУБЛІКАТІВ УВІМКНЕНА ===")
            duplicates, dup_stats = processor.check_duplicates()
            
            if dup_stats.get('total', 0) > 0:
                print(f"   ⚠️ Знайдено дублікатів: {dup_stats['total']}")
                for sheet, count in dup_stats.get('by_sheet', {}).items():
                    if count > 0:
                        print(f"      {sheet}: {count}")
            else:
                print(f"   ✅ Дублікатів не знайдено")
        else:
            print("\n   ⚠️ Параметр enable_duplicates = False, пропускаємо перевірку")
        
        # Пересоздаємо лист LOG з результатами перевірок
        processor.create_log_sheet()
        
        # Зберігаємо фінальний результат з логами перевірок
        processor.save()
        
        print(f"\n✅ === ОБРОБКА ЗАВЕРШЕНА ===")
        print(f"📊 Всього скопійовано: {total_copied} рядків")
        
        # Повертаємо результат
        result = {
            'success': True,
            'total_rows': total_copied
        }
        print(f"\n__RESULT__{json.dumps(result)}__END__")
        
    except Exception as e:
        print(f"\n❌ ПОМИЛКА: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        
        result = {
            'success': False,
            'error': str(e)
        }
        print(f"\n__RESULT__{json.dumps(result)}__END__")
        sys.exit(1)


if __name__ == '__main__':
    main()
