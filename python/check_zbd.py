"""
check_zbd.py - Перевірка табелю на відповідність даним ЖБД

Функціонал:
- Читає Word файли ЖБД (декілька файлів) - витягує дані про прибуття/вибуття
- Читає Excel табель (колонка C - ПІБ, E1:AI1 - дати)
- Читає конфігураційний Excel (колонка B - позиції, C - виплати)
- Перевіряє відповідність виплат в табелі згідно даних ЖБД
- Позначає червоним помилки: ПІБ не знайдено або виплата невірна
- Створює результат у форматі Excel
"""

import sys
import io
import json
import re
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime, date, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from docx import Document
import builtins

_log_file_handle = None


def _set_log_file(path: Optional[str]) -> None:
    """Встановлює шлях для лог-файлу і відкриває його для запису."""
    global _log_file_handle

    if _log_file_handle:
        try:
            _log_file_handle.close()
        finally:
            _log_file_handle = None

    if path:
        log_path = Path(path)
        if not log_path.parent.exists():
            log_path.parent.mkdir(parents=True, exist_ok=True)
        _log_file_handle = log_path.open('w', encoding='utf-8')


def _log_print(*args, **kwargs) -> None:
    """Друкує повідомлення в консоль та дублює у лог-файл."""
    builtins.print(*args, **kwargs)

    if _log_file_handle:
        sep = kwargs.get('sep', ' ')
        end = kwargs.get('end', '\n')
        text = sep.join(str(arg) for arg in args)
        _log_file_handle.write(text + end)
        _log_file_handle.flush()


print = _log_print

# Встановлюємо UTF-8 для stdout/stderr на Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


class ZBDChecker:
    """Перевірка табелю на відповідність даним ЖБД"""

    def __init__(self, config_excel: str = None):
        self.errors = []
        self.warnings = []
        self.config_excel = config_excel
        
        # Дані з конфігурації: {позиція: виплата}
        self.position_payments: Dict[str, str] = {}
        
        # Дані з ЖБД: {ПІБ зі званням: [(позиція, блок, дата_прибуття, дата_вибуття)]}
        self.zbd_data: Dict[str, List[Tuple[str, str, Optional[date], Optional[date]]]] = {}
        
        # Індекс для пошуку по ПІБ без звання: {ПІБ без звання: ПІБ зі званням}
        self.pib_index: Dict[str, str] = {}
        
        # Місяць для перевірки
        self.check_month: Optional[int] = None
        self.check_year: Optional[int] = None
        
        # Список звань для видалення
        self.ranks = [
            'солдат', 'старший солдат',
            'капрал', 'молодший капрал',
            'молодший сержант', 'сержант', 'старший сержант', 'головний сержант', 'майстер-сержант', 'штаб-сержант',
            'прапорщик', 'старший прапорщик',
            'молодший лейтенант', 'лейтенант', 'старший лейтенант', 'капітан',
            'майор', 'підполковник', 'полковник',
            'генерал-майор', 'генерал-лейтенант', 'генерал-полковник', 'генерал армії України'
        ]

    def _normalize_pib(self, pib: str) -> str:
        """
        Нормалізує ПІБ (видаляє звання)
        
        Args:
            pib: ПІБ зі званням або без
            
        Returns:
            ПІБ без звання
        """
        # Замінюємо табуляції на пробіли та видаляємо зайві пробіли
        pib_clean = re.sub(r'\s+', ' ', pib).strip()
        
        # Сортуємо звання від довшого до коротшого (щоб "молодший сержант" оброблявся перед "сержант")
        sorted_ranks = sorted(self.ranks, key=len, reverse=True)
        
        # Видаляємо звання з початку
        for rank in sorted_ranks:
            if pib_clean.lower().startswith(rank.lower()):
                pib_clean = pib_clean[len(rank):].strip()
                break
        
        return pib_clean

    @staticmethod
    def _normalize_payment_value(value: Optional[Any]) -> Optional[str]:
        """Переводить значення виплати у стандартизований вигляд (кирилиця, верхній регістр)."""
        if value is None:
            return None

        text = str(value).strip()
        if not text:
            return None

        mapping = {
            'A': 'А', 'B': 'В', 'C': 'С', 'E': 'Е', 'H': 'Н', 'K': 'К', 'M': 'М', 'O': 'О',
            'P': 'Р', 'R': 'Р', 'T': 'Т', 'X': 'Х', 'Y': 'У', 'V': 'В', 'Z': 'З', '3': 'З',
            'a': 'а', 'b': 'в', 'c': 'с', 'e': 'е', 'h': 'н', 'k': 'к', 'm': 'м', 'o': 'о',
            'p': 'р', 'r': 'р', 't': 'т', 'x': 'х', 'y': 'у', 'v': 'в', 'z': 'з'
        }

        normalized_chars = [mapping.get(ch, ch) for ch in text]
        normalized = ''.join(normalized_chars).upper()

        return normalized

    def check_files(self, word_files: List[str], excel_file: str, output_path: str) -> Dict[str, Any]:
        """
        Перевіряє Word та Excel файли

        Args:
            word_files: Список шляхів до Word файлів ЖБД
            excel_file: Шлях до Excel табелю
            output_path: Шлях для збереження звіту

        Returns:
            Словник з результатами перевірки
        """
        print(f"📝 Перевірка {len(word_files)} Word файлів...")
        
        # Перевіряємо конфігурацію
        if self.config_excel:
            print(f"⚙️ Використання конфігураційного Excel: {Path(self.config_excel).name}")
            self._load_config()
        
        # Перевіряємо Word файли (витягуємо дані ЖБД)
        for i, word_file in enumerate(word_files, 1):
            print(f"  {i}. {Path(word_file).name}")
            self._parse_zbd_word(word_file)
        
        print(f"\n📖 Читання табелю Excel: {excel_file}")

        try:
            # Читаємо табель
            wb_tabel = load_workbook(excel_file, data_only=True)
            
            # Перевіряємо табель згідно даних ЖБД
            self._check_tabel(wb_tabel, excel_file, output_path)
            
            wb_tabel.close()
            
            print(f"\n✅ Перевірка завершена")
            print(f"  - Помилок: {len(self.errors)}")
            print(f"  - Попереджень: {len(self.warnings)}")
            
            return {
                'success': True,
                'errors': len(self.errors),
                'warnings': len(self.warnings)
            }

        except Exception as e:
            print(f"❌ Помилка: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': str(e)
            }

    def _load_config(self) -> None:
        """Завантажує конфігурацію позицій та виплат"""
        print(f"⚙️ Завантаження конфігурації...")
        
        try:
            wb = load_workbook(self.config_excel, data_only=True)
            
            # Шукаємо аркуш РВП
            if 'РВП' in wb.sheetnames:
                ws = wb['РВП']
            else:
                ws = wb.active
            
            # Читаємо дані: колонка B - позиції, C - виплати
            for row in ws.iter_rows(min_row=2, values_only=True):
                position = row[1]  # Колонка B
                payment = row[2]   # Колонка C
                
                if position and payment:
                    position_str = str(position).strip()
                    payment_str = str(payment).strip()
                    self.position_payments[position_str] = payment_str
                    
                    # Логування для "обстріл"
                    if 'обстріл' in position_str.lower():
                        print(f"  💥 Знайдено конфігурацію для обстрілу: '{position_str}' → '{payment_str}'")
            
            wb.close()
            print(f"  ✅ Завантажено {len(self.position_payments)} позицій")
            
        except Exception as e:
            print(f"  ⚠️ Помилка завантаження конфігурації: {e}")

    def _parse_zbd_word(self, file_path: str) -> None:
        """Читає Word файл ЖБД та витягує дані про прибуття/вибуття"""
        try:
            if not Path(file_path).exists():
                self.errors.append(f"Word файл не знайдено: {Path(file_path).name}")
                return
            
            # Спробуємо визначити місяць з назви файлу
            filename = Path(file_path).stem.lower()
            month_names = {
                'січень': 1, 'січня': 1,
                'лютий': 2, 'лютого': 2,
                'березень': 3, 'березня': 3,
                'квітень': 4, 'квітня': 4,
                'травень': 5, 'травня': 5,
                'червень': 6, 'червня': 6,
                'липень': 7, 'липня': 7,
                'серпень': 8, 'серпня': 8,
                'вересень': 9, 'вересня': 9,
                'жовтень': 10, 'жовтня': 10,
                'листопад': 11, 'листопада': 11,
                'грудень': 12, 'грудня': 12
            }
            
            # Шукаємо назву місяця в імені файлу
            for month_name, month_num in month_names.items():
                if month_name in filename:
                    if not hasattr(self, 'check_month'):
                        self.check_month = month_num
                        # Якщо рік не встановлено, використовуємо поточний
                        if not hasattr(self, 'check_year'):
                            from datetime import datetime
                            self.check_year = datetime.now().year
                        print(f"  📅 Визначено місяць з назви файлу: {month_num} ({month_name.capitalize()}), рік: {self.check_year}")
                    break
            
            print(f"  📄 Парсинг {Path(file_path).name}...")
            doc = Document(file_path)
            people_found = 0
            
            # ЖБД складається з 31 таблиці (по одній на кожен день місяця)
            for table_idx, table in enumerate(doc.tables, 1):
                # Логування тільки кожної 5-ої таблиці
                if table_idx % 5 == 1:
                    print(f"    🔍 Обробка таблиць {table_idx}-{min(table_idx+4, 31)}...")
                
                current_position = None
                current_block = None  # 'прибув' або 'вибув'
                
                # Витягуємо дату з рядка 2, колонка 0 (формат 01.09.2025)
                event_date = None
                if len(table.rows) > 2 and len(table.rows[2].cells) > 0:
                    date_text = table.rows[2].cells[0].text.strip()
                    event_date = self._parse_date(date_text)
                    if event_date:
                        # Логування тільки для першої таблиці
                        if table_idx == 1:
                            print(f"      📅 Перша дата: {event_date.strftime('%d.%m.%Y')}")
                        # Встановлюємо рік і місяць для перевірки при першій знайденій даті
                        if self.check_year is None or self.check_month is None:
                            self.check_year = event_date.year
                            self.check_month = event_date.month
                            print(f"      ✓ Встановлено період перевірки: {self.check_month}/{self.check_year}")
                    else:
                        print(f"      ⚠️ Не вдалося розпарсити дату '{date_text}'")
                
                if not event_date:
                    print(f"      ❌ Пропуск таблиці - не знайдено дату")
                    continue
                
                for row_idx, row in enumerate(table.rows):
                    cells = row.cells
                    if len(cells) < 2:
                        continue
                    
                    # Текст з колонки "Завдання військ..." (індекс 1)
                    cell_text = cells[1].text.strip()
                    
                    if not cell_text:
                        continue
                    
                    # Логування перших 5 рядків першої таблиці для діагностики
                    if table_idx == 1 and row_idx <= 5:
                        print(f"      [DEBUG] Рядок {row_idx}: '{cell_text[:80]}'...")
                    
                    # Шукаємо заголовок з позицією (наприклад, "Проведено ротацію о/с ВП "Маріо-1"")
                    if self._is_position_header(cell_text):
                        current_position = self._extract_position(cell_text)
                        # Логування тільки для обстрілу
                        if current_position == 'обстріл':
                            print(f"      💥 Обстріл, таблиця {table_idx}")
                        # НЕ робимо continue - перевіряємо чи є прибули/вибули в тому ж рядку
                    
                    # Для "обстріл" - окрема логіка (шукаємо "перебували:")
                    if current_position == 'обстріл' and 'перебували:' in cell_text.lower():
                        # Парсимо ПІБ після "перебували:" (всі отримують виплату за цей день)
                        result = self._parse_person_row_from_text(event_date, cell_text, current_position, 'обстріл')
                        if result > 0:
                            people_found += result
                        continue
                    
                    # Шукаємо блок прибули:/вибули: (може бути в тому ж рядку що і позиція)
                    if 'прибули:' in cell_text.lower():
                        current_block = 'прибув'
                        
                        # Парсимо ПІБ з цього ж рядка
                        if current_position:
                            result = self._parse_person_row_from_text(event_date, cell_text, current_position, current_block)
                            if result > 0:
                                people_found += result
                    
                    if 'вибули:' in cell_text.lower():
                        current_block = 'вибув'
                        
                        # Парсимо ПІБ з цього ж рядка
                        if current_position:
                            result = self._parse_person_row_from_text(event_date, cell_text, current_position, current_block)
                            if result > 0:
                                people_found += result

                    if 'штурман:' in cell_text.lower():
                        current_block = 'штурман'

                        if current_position:
                            result = self._parse_person_row_from_text(event_date, cell_text, current_position, current_block)
                            if result > 0:
                                people_found += result
            
        except Exception as e:
            self.errors.append(f"Помилка читання Word файлу {Path(file_path).name}: {str(e)}")
            import traceback
            traceback.print_exc()

    def _is_position_header(self, text: str) -> bool:
        """Перевіряє, чи є це заголовок з позицією"""
        # Шукаємо назву позиції в конфігурації (будь-яку)
        text_lower = text.lower()
        for position in self.position_payments.keys():
            if position.lower() in text_lower:
                return True
        return False

    def _extract_position(self, text: str) -> str:
        """Витягує назву позиції з заголовка"""
        text_lower = text.lower()
        for position in self.position_payments.keys():
            if position.lower() in text_lower:
                return position
        return text.strip()

    def _parse_person_row_from_text(self, event_date: date, content_text: str, position: str, block: str) -> int:
        """
        Парсить ПІБ з тексту комірки Word.
        
        Args:
            event_date: Дата події (з таблиці Word)
            content_text: Текст з колонки "Завдання військ..."
            position: Позиція
            block: Блок (прибув/вибув/обстріл)
        
        Returns:
            Кількість знайдених ПІБ
        """
        found_count = 0
        
        try:
            if not hasattr(self, 'check_year') or not hasattr(self, 'check_month'):
                self.check_year = event_date.year
                self.check_month = event_date.month

            lines = content_text.split('\n')

            if block == 'обстріл':
                target_marker = 'перебували:'
            elif block == 'прибув':
                target_marker = 'прибули:'
            elif block == 'штурман':
                target_marker = 'штурман:'
            else:
                target_marker = 'вибули:'

            collecting = False
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                normalized = line.lower()

                if target_marker in normalized:
                    collecting = True
                    continue

                if collecting and (
                    'прибули:' in normalized or
                    'вибули:' in normalized or
                    'водій:' in normalized or
                    'штурман:' in normalized
                ):
                    break

                if not collecting:
                    continue

                if 'проведено' in normalized or 'ротац' in normalized:
                    continue

                if len(line) < 5:
                    continue

                pib = line

                if block in ('обстріл', 'штурман'):
                    date_arrival = event_date
                    date_departure = event_date
                else:
                    date_arrival = event_date if block == 'прибув' else None
                    date_departure = event_date if block == 'вибув' else None

                if pib not in self.zbd_data:
                    self.zbd_data[pib] = []

                self.zbd_data[pib].append((position, block, date_arrival, date_departure))

                pib_normalized = self._normalize_pib(pib)
                self.pib_index[pib_normalized] = pib

                found_count += 1

                if found_count == 1:
                    print(f"        → ПІБ приклад: '{pib_normalized}'")

            return found_count

        except Exception:
            return 0

    def _parse_person_row(self, cells, position: str, block: str) -> bool:
        """
        Парсить рядок з даними про людину з таблиці Word.
        Структура таблиці в Word:
        - Колонка 0: Дата, час (пункт 1)
        - Колонка 1: Завдання військ... (пункт 2) - тут знаходяться прибули:/вибули: та ПІБ
        - Колонка 2: Примітка (пункт 3)
        
        Повертає True якщо успішно
        """
        try:
            if len(cells) < 2:
                return False
            
            # ПІБ знаходиться у другій колонці (індекс 1) - "Завдання військ..."
            # В цій колонці є багато рядків, шукаємо ПІБ (після прибули:/вибули:)
            cell_text = cells[1].text.strip()
            
            if not cell_text:
                return False
            
            # Розбиваємо текст на рядки
            lines = cell_text.split('\n')
            
            # Пропускаємо заголовки, шукаємо ПІБ
            found_block_marker = False
            for line in lines:
                line = line.strip()
                
                # Пропускаємо порожні рядки та заголовки
                if not line or 'проведено' in line.lower() or 'ротац' in line.lower():
                    continue
                
                # Якщо знайшли маркер блоку
                if 'прибули:' in line.lower() or 'вибули:' in line.lower():
                    found_block_marker = True
                    continue
                
                # Якщо ми після маркера блоку і це схоже на ПІБ
                if found_block_marker and len(line) > 5:
                    # Перевіряємо, що це не інший маркер
                    if 'водій:' in line.lower() or 'штурман:' in line.lower():
                        break
                    
                    # Це ПІБ - додаємо
                    pib = line
                    
                    # Шукаємо дати в колонці "Дата, час" (індекс 0)
                    date_text = cells[0].text.strip() if len(cells) > 0 else ''
                    
                    # Витягуємо тільки дату (без часу)
                    date_match = re.search(r'(\d{1,2}\.\d{1,2}\.\d{4})', date_text)
                    event_date = self._parse_date(date_match.group(1)) if date_match else None
                    
                    # Для блоку "прибув" - це дата прибуття
                    # Для блоку "вибув" - це дата вибуття
                    date_arrival = event_date if block == 'прибув' else None
                    date_departure = event_date if block == 'вибув' else None
                    
                    # Додаємо до даних
                    if pib not in self.zbd_data:
                        self.zbd_data[pib] = []
                    
                    self.zbd_data[pib].append((position, block, date_arrival, date_departure))
                    return True
            
            return False
            
        except Exception as e:
            return False  # Пропускаємо некоректні рядки

    def _parse_date(self, date_str: str) -> Optional[date]:
        """Парсить дату з різних форматів"""
        if not date_str or date_str == '-':
            return None
        
        # Прибираємо зайві символи
        date_str = date_str.strip().replace('\n', ' ')
        
        # Формати: DD.MM.YYYY, DD.MM.YY
        patterns = [
            r'(\d{1,2})\.(\d{1,2})\.(\d{4})',
            r'(\d{1,2})\.(\d{1,2})\.(\d{2})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, date_str)
            if match:
                day = int(match.group(1))
                month = int(match.group(2))
                year = int(match.group(3))
                
                if year < 100:
                    year += 2000
                
                try:
                    return date(year, month, day)
                except:
                    pass
        
        return None

    def _check_tabel(self, wb: Workbook, excel_file: str, output_path: str) -> None:
        """Перевіряє табель згідно даних ЖБД"""
        ws = wb.active
        
        # Читаємо дати з E1:AI1
        dates = []
        for col in range(5, 36):  # E до AI (5 до 35)
            cell_value = ws.cell(1, col).value
            if cell_value:
                # СПОЧАТКУ перевіряємо чи це номер дня (1-31)
                parsed_date = None
                try:
                    day_number = int(cell_value)
                    if 1 <= day_number <= 31 and self.check_year and self.check_month:
                        # Створюємо дату з номера дня та встановленого року/місяця
                        parsed_date = date(self.check_year, self.check_month, day_number)
                        dates.append((col, parsed_date))
                        if len(dates) <= 3:  # Логування для перших 3 дат
                            print(f"  [DEBUG] Колонка {col}: день {day_number} → {parsed_date.strftime('%d.%m.%Y')}")
                except (ValueError, TypeError):
                    # Якщо не вдалося як число, пробуємо парсити як дату
                    parsed_date = self._parse_date_header(cell_value)
                    if parsed_date:
                        dates.append((col, parsed_date))
        
        print(f"  📅 Знайдено {len(dates)} дат у табелі")
        print(f"  📅 Місяць перевірки: {self.check_month}/{self.check_year}")
        print(f"  👥 Знайдено {len(self.zbd_data)} ПІБ у ЖБД")
        
        # Виводимо перші 5 ПІБ для перевірки
        if self.zbd_data:
            print(f"  📋 Приклади знайдених ПІБ:")
            for i, pib in enumerate(list(self.zbd_data.keys())[:5], 1):
                print(f"    {i}. {pib}")
        
        # Виводимо перші 5 нормалізованих ПІБ з індексу
        if self.pib_index:
            print(f"  📋 Приклади нормалізованих ПІБ (індекс):")
            for i, (norm_pib, orig_pib) in enumerate(list(self.pib_index.items())[:5], 1):
                print(f"    {i}. '{norm_pib}' → '{orig_pib}'")
        
        # Створюємо копію для результату
        wb_result = load_workbook(excel_file, data_only=False)
        ws_result = wb_result.active
        
        # Перевіряємо кожен ПІБ у табелі
        row_num = 2
        people_checked = 0
        people_not_found = 0
        
        while True:
            pib_cell = ws.cell(row_num, 3)  # Колонка C
            pib = pib_cell.value
            
            if not pib:
                break
            
            pib = str(pib).strip()
            
            # Логування для першої людини
            if people_checked == 0:
                print(f"\n  🔍 Перевірка першого ПІБ: '{pib}'")
            
            # Спробуємо знайти ПІБ: спочатку точний пошук, потім нормалізований
            found_pib = None
            if pib in self.zbd_data:
                found_pib = pib
                if people_checked == 0:
                    print(f"    ✓ Знайдено (точний збіг)")
            else:
                # Пробуємо нормалізований пошук
                pib_normalized = self._normalize_pib(pib)
                if pib_normalized in self.pib_index:
                    found_pib = self.pib_index[pib_normalized]
                    if people_checked == 0:
                        print(f"    ✓ Знайдено (нормалізація)")
                else:
                    if people_checked == 0:
                        print(f"    ✗ НЕ знайдено в ЖБД")
            
            if found_pib is None:
                people_not_found += 1
            
            self._check_person_in_tabel(ws, ws_result, row_num, pib, dates, found_pib)
            people_checked += 1
            row_num += 1
        
        print(f"\n  📊 Перевірено {people_checked} ПІБ")
        print(f"  ❌ Не знайдено в ЖБД: {people_not_found}")
        
        # Зберігаємо результат
        wb_result.save(output_path)
        wb_result.close()
        print(f"  💾 Результат збережено: {output_path}")

    def _parse_date_header(self, value) -> Optional[date]:
        """Парсить дату із заголовка табелю"""
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, date):
            return value
        elif isinstance(value, (int, float)):
            # Excel date number
            try:
                return datetime(1899, 12, 30) + timedelta(days=int(value))
            except:
                pass
        elif isinstance(value, str):
            return self._parse_date(value)
        
        return None

    def _check_person_in_tabel(self, ws, ws_result, row_num: int, pib: str, dates: List[Tuple[int, date]], found_pib: Optional[str]) -> None:
        """Перевіряє одну людину в табелі
        
        Args:
            ws: Оригінальний аркуш
            ws_result: Аркуш результату
            row_num: Номер рядка
            pib: ПІБ з табелю (без звання)
            dates: Список дат із заголовків
            found_pib: ПІБ знайдений в ЖБД (зі званням) або None якщо не знайдено
        """
        
        # Перевіряємо, чи є ПІБ у ЖБД
        if found_pib is None:
            # ПІБ не знайдено - позначаємо червоним
            ws_result.cell(row_num, 3).fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            self.errors.append(f"ПІБ не знайдено в ЖБД: {pib}")
            return
        
        # ПІБ знайдено - позначаємо зеленим
        ws_result.cell(row_num, 3).fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        
        # Отримуємо дані з ЖБД (використовуємо found_pib зі званням)
        person_records = self.zbd_data[found_pib]
        
        # Запам'ятовуємо поточного ПІБ для дебагу (використовується в _get_expected_payment)
        self._current_pib = pib

        # Перевіряємо кожну дату в табелі
        first_date_logged = False
        for col, check_date in dates:
            expected_payment_raw = self._get_expected_payment(person_records, check_date)
            expected_payment = self._normalize_payment_value(expected_payment_raw)

            actual_cell_value = ws.cell(row_num, col).value
            actual_payment_raw = str(actual_cell_value).strip() if actual_cell_value is not None else None
            actual_payment = self._normalize_payment_value(actual_cell_value)

            # Логування для першої дати першої людини
            if not first_date_logged and row_num == 2:
                print(f"    [DEBUG] Пошук виплати: ПІБ рядок {row_num}, дата {check_date.strftime('%d.%m.%Y')}, колонка {col}")
                print(f"    [DEBUG] Очікувана виплата (raw): '{expected_payment_raw}' → нормалізовано: '{expected_payment}'")
                print(f"    [DEBUG] Фактична виплата в табелі (raw): '{actual_payment_raw}' → нормалізовано: '{actual_payment}'")
                print(f"    [DEBUG] Збіг: {expected_payment == actual_payment}")
                first_date_logged = True

            # Пропускаємо перевірку тільки якщо очікується Р або ЗВР (залишаємо біле)
            if expected_payment in ['Р', 'ЗВР']:
                continue
            
            # Порівнюємо
            if expected_payment != actual_payment:
                if expected_payment is None:
                    if actual_payment is None:
                        ws_result.cell(row_num, col).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                        continue

                    if actual_payment in ['Р', 'ЗВР']:
                        ws_result.cell(row_num, col).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                        continue

                    if actual_payment not in ['Б', 'Н']:
                        ws_result.cell(row_num, col).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                        continue

                # Невірна виплата - позначаємо червоним
                ws_result.cell(row_num, col).fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                expected_err = expected_payment if expected_payment is not None else expected_payment_raw
                actual_err = actual_payment if actual_payment is not None else actual_payment_raw
                self.errors.append(f"{pib}, дата {check_date}: очікувалось '{expected_err}', знайдено '{actual_err}'")
            else:
                # Вірна виплата - позначаємо зеленим
                if expected_payment is not None:  # Позначаємо тільки якщо має бути виплата
                    ws_result.cell(row_num, col).fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

        # Скидаємо дебаг-поле
        self._current_pib = None

    def _get_expected_payment(self, records: List[Tuple[str, str, Optional[date], Optional[date]]], check_date: date) -> Optional[str]:
        """Визначає очікувану виплату для конкретної дати"""

        if isinstance(check_date, datetime):
            check_date = check_date.date()

        if self.check_month is None or self.check_year is None:
            self.check_year = check_date.year
            self.check_month = check_date.month

        current_pib = getattr(self, '_current_pib', 'невідомо')
        print(f"\n    [DEBUG] === Пошук виплати для '{current_pib}' на {check_date.strftime('%d.%m.%Y')} ===")
        print(f"    [DEBUG] Записів у ЖБД для цієї людини: {len(records)}")

        def _log_record(idx: int, position: str, block: str, date_arrival: Optional[date], date_departure: Optional[date]) -> None:
            print(f"    [DEBUG] Запис #{idx+1}: позиція='{position}', блок='{block}'")
            print(f"    [DEBUG]   Прибуття: {date_arrival.strftime('%d.%m.%Y') if isinstance(date_arrival, date) else 'не вказано'}")
            print(f"    [DEBUG]   Вибуття: {date_departure.strftime('%d.%m.%Y') if isinstance(date_departure, date) else 'не вказано'}")

        departures_by_position: Dict[str, List[date]] = {}
        arrivals_by_position: Dict[str, List[date]] = {}
        for position, block, _date_arrival, date_departure in records:
            if block == 'вибув' and isinstance(date_departure, date):
                departures_by_position.setdefault(position, []).append(date_departure)
            if block == 'прибув' and isinstance(_date_arrival, date):
                arrivals_by_position.setdefault(position, []).append(_date_arrival)

        for pos in departures_by_position:
            departures_by_position[pos].sort()
        for pos in arrivals_by_position:
            arrivals_by_position[pos].sort()

        # Спочатку перевіряємо всі записи обстрілів, щоб вони мали найвищий пріоритет
        priority_blocks = ('обстріл', 'штурман')

        for priority in priority_blocks:
            for idx, (position, block, date_arrival, date_departure) in enumerate(records):
                if block != priority:
                    continue

                _log_record(idx, position, block, date_arrival, date_departure)

                if priority == 'обстріл':
                    print(f"    [DEBUG]   💥 Обстріл: дата події {date_arrival.strftime('%d.%m.%Y') if isinstance(date_arrival, date) else 'невідома'}")
                    print(f"    [DEBUG]      Перевіряємо дату {check_date.strftime('%d.%m.%Y')}")

                    if date_arrival and date_arrival == check_date:
                        payment = self._lookup_payment(position)
                        print(f"    [DEBUG]   → ✓ Збіг дат, повертаємо виплату '{payment}'")
                        return payment

                    print(f"    [DEBUG]   → ✗ Інша дата, переходимо до наступного запису")
                else:
                    one_day_event = date_arrival or date_departure
                    print(f"    [DEBUG]   🧭 Штурман: дата події {one_day_event.strftime('%d.%m.%Y') if isinstance(one_day_event, date) else 'невідома'}")
                    print(f"    [DEBUG]      Перевіряємо дату {check_date.strftime('%d.%m.%Y')}")

                    if one_day_event and one_day_event == check_date:
                        payment = self._lookup_payment(position)
                        print(f"    [DEBUG]   → ✓ Одноденний запис, повертаємо виплату '{payment}'")
                        return payment

                    print(f"    [DEBUG]   → ✗ Інша дата, переходимо до наступного запису")

        # Далі перевіряємо інші записи (прибуття/вибуття)
        for idx, (position, block, date_arrival, date_departure) in enumerate(records):
            if block in priority_blocks:
                continue

            _log_record(idx, position, block, date_arrival, date_departure)

            start_date = date_arrival if date_arrival else date(self.check_year, self.check_month, 1)
            end_date = date_departure if date_departure else self._get_last_day_of_month()

            if block == 'прибув' and isinstance(date_arrival, date):
                possible_departures = [dep for dep in departures_by_position.get(position, []) if dep >= date_arrival]
                if possible_departures:
                    matched_departure = possible_departures[0]
                    if matched_departure < end_date:
                        print(f"    [DEBUG]   → Знайдено наступний 'вибув' {matched_departure.strftime('%d.%m.%Y')} для позиції '{position}', обмежуємо період")
                        end_date = matched_departure

            if block == 'вибув' and not date_arrival:
                if isinstance(date_departure, date):
                    possible_arrivals = [arr for arr in arrivals_by_position.get(position, []) if arr <= date_departure]
                    if possible_arrivals:
                        inferred_arrival = possible_arrivals[-1]
                        print(f"    [DEBUG]   ⚠️ 'Вибув' без прибуття → використовуємо останній прибув {inferred_arrival.strftime('%d.%m.%Y')} для позиції '{position}'")
                        start_date = inferred_arrival
                    else:
                        print(f"    [DEBUG]   ⚠️ 'Вибув' без прибуття → старт з початку місяця")
                else:
                    print(f"    [DEBUG]   ⚠️ 'Вибув' без прибуття → старт з початку місяця")
            if block == 'прибув' and not date_departure:
                print(f"    [DEBUG]   ℹ️ 'Прибув' без вибуття → до кінця місяця")

            month_start = date(self.check_year, self.check_month, 1)
            month_end = self._get_last_day_of_month()
            if start_date < month_start:
                start_date = month_start
            if end_date > month_end:
                end_date = month_end

            print(f"    [DEBUG]   → Період для перевірки: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
            print(f"    [DEBUG]   → Дата в періоді? {start_date <= check_date <= end_date}")

            if start_date <= check_date <= end_date:
                payment = self._lookup_payment(position)
                print(f"    [DEBUG]   → ✓ Дата всередині періоду, виплата '{payment}'")
                if payment is None:
                    print(f"    [DEBUG]   → ⚠️ Позиція '{position}' не знайдена у конфігурації")
                    print(f"    [DEBUG]   → Доступні позиції: {list(self.position_payments.keys())[:5]}")
                return payment

            print(f"    [DEBUG]   → ✗ Дата не входить у період")

        print(f"    [DEBUG] === Результат: виплати немає (None) ===\n")
        return None

    def _lookup_payment(self, position: Optional[str]) -> Optional[str]:
        """Повертає виплату для позиції, враховуючи можливі варіації ключів"""

        if not position:
            return None

        candidates = [position, position.strip(), position.lower(), position.upper()]
        for key in candidates:
            if key and key in self.position_payments:
                return self.position_payments[key]

        return None

    def _get_last_day_of_month(self) -> date:
        """Повертає останній день місяця перевірки"""
        from calendar import monthrange
        from datetime import datetime
        
        # Якщо місяць не встановлено, використовуємо поточний
        if self.check_month is None or self.check_year is None:
            now = datetime.now()
            check_year = self.check_year if self.check_year else now.year
            check_month = self.check_month if self.check_month else now.month
        else:
            check_year = self.check_year
            check_month = self.check_month
            
        last_day = monthrange(check_year, check_month)[1]
        return date(check_year, check_month, last_day)


def main():
    """Головна функція"""
    try:
        # Читаємо конфігурацію через stdin
        config_json = sys.stdin.read()
        config = json.loads(config_json)

        word_files = config.get('word_files', [])
        excel_file = config.get('excel_file', '')
        config_excel = config.get('config_excel', None)
        output_file = config.get('output_file', '')
        log_file = config.get('log_file')

        if not word_files:
            print("❌ Не вказано Word файли")
            sys.exit(1)

        if not excel_file:
            print("❌ Не вказано Excel файл")
            sys.exit(1)

        if not output_file:
            print("❌ Не вказано файл для збереження результатів")
            sys.exit(1)

        # Перевіряємо чи існує Excel файл
        if not Path(excel_file).exists():
            print(f"❌ Excel файл не знайдено: {excel_file}")
            sys.exit(1)

        # Перевіряємо чи існують Word файли
        for word_file in word_files:
            if not Path(word_file).exists():
                print(f"❌ Word файл не знайдено: {word_file}")
                sys.exit(1)

        # Перевіряємо чи існує конфігураційний Excel (якщо вказано)
        if config_excel and not Path(config_excel).exists():
            print(f"❌ Конфігураційний Excel не знайдено: {config_excel}")
            sys.exit(1)

        if not log_file:
            log_file = str(Path(output_file).with_suffix('.log'))

        _set_log_file(log_file)
        print(f"ℹ️ Логи дублюються у файл: {log_file}")

        # Створюємо checker
        checker = ZBDChecker(config_excel=config_excel)
        
        # Виконуємо перевірку
        result = checker.check_files(word_files, excel_file, output_file)
        
        if not result.get('success'):
            sys.exit(1)

    except json.JSONDecodeError as e:
        print(f"❌ Помилка парсингу JSON конфігурації: {str(e)}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Неочікувана помилка: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        _set_log_file(None)


if __name__ == '__main__':
    main()
