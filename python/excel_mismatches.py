"""
Excel Mismatches Checker - Порівняння листів ЗС та БЗ
Адаптація VBA логіки для пошуку невідповідностей

Правила порівняння:
- ЗС: колонка D (ключ ПІБ), фільтр F="БЗ", ігнор токенів у D (ПІБ)
  Структура ЗС: B=підрозділ, C=звання, D=ПІБ, E=псевдо, F=статус, G=дод.
- БЗ: колонка E (ключ ПІБ), без фільтра F, ігнор токенів у E (ПІБ), ігнор H містить "придані", ліміт 898
  Структура БЗ: C=підрозділ, D=звання, E=ПІБ, F=псевдо, G=статус, H=розташування
"""

from typing import Dict, List, Tuple, Optional, Set
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import re

# ========= НАЛАШТУВАННЯ =========
S1_NAME = "ЗС"   # лист 1 у цільовій книзі
S2_NAME = "БЗ"   # лист 2 у цільовій книзі

# ЗС: B=підрозділ, C=звання, D=ПІБ, E=псевдо, F=статус, G=дод.
S1_COL_KEY = 4   # D - ПІБ (1-indexed)
S1_COL_F = 6     # F - статус (фільтр "БЗ")
S1_COL_G = 7     # G - дод.

# БЗ: C=підрозділ, D=звання, E=ПІБ, F=псевдо, G=статус, H=розташування
S2_COL_KEY = 5   # E - ПІБ
S2_COL_H = 8     # H - розташування (для ігнору "придані")

F_REQUIRED = "БЗ"          # що має бути у колонці F на ЗС
MAX_ROW_LIMIT = 898        # ЛІМІТ РЯДКІВ — ЛИШЕ ДЛЯ ЛИСТА "БЗ"
H_IGNORE_TOKEN = "придані" # ІГНОР у H — ЛИШЕ ДЛЯ ЛИСТА "БЗ"
PREVIEW_MAX = 25

# Кого ігноруємо (за ключем-рядком у колонці ключів)
IGNORE_TOKENS = [
    "1РСпП", "2РСпП", "3РСпП", "РВПСпП", "РМТЗ", "Медичний пункт",
    "Взвод зв'язку", "Відділення інструкторів", "ВРЕБ", "ВРСП",
    "М<", "РБПС",
    "Мінометна батарея", "Рота безпілотних систем", "Взвод РЕБ",
    "Нерозподілені", "Придані",
    "1р СпП"  # Варіант з пробілом та малою літерою
]


class MismatchEntry:
    """Запис про невідповідність"""
    def __init__(self, sheet: str, cell_addr: str, value: str, reason: str):
        self.sheet = sheet
        self.cell_addr = cell_addr
        self.value = value
        self.reason = reason
    
    def __str__(self) -> str:
        return f"{self.sheet}!{self.cell_addr} = «{self.value}» — {self.reason}"


class ExcelMismatchChecker:
    """Перевірка невідповідностей між листами ЗС та БЗ"""
    
    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.wb = None
        self.ws1 = None  # ЗС
        self.ws2 = None  # БЗ
    
    def check(self) -> Tuple[List[MismatchEntry], str]:
        """
        Перевірити невідповідності
        
        Returns:
            (список_невідповідностей, повідомлення_про_помилку)
        """
        try:
            # Завантажуємо книгу
            self.wb = load_workbook(self.workbook_path, data_only=True)
            
            # Шукаємо листи
            self.ws1 = self._find_sheet(S1_NAME)
            self.ws2 = self._find_sheet(S2_NAME)
            
            if self.ws1 is None:
                return [], f"Не знайдено аркуш '{S1_NAME}'"
            if self.ws2 is None:
                return [], f"Не знайдено аркуш '{S2_NAME}'"
            
            # Визначаємо межі даних
            r1_first, r1_last = 4, self._max_last_row(self.ws1, [S1_COL_KEY, S1_COL_F, S1_COL_G])
            r2_first, r2_last = 4, self._max_last_row(self.ws2, [S2_COL_KEY, S2_COL_H])
            
            # БЗ: зрізаємо до MAX_ROW_LIMIT
            if r2_last > MAX_ROW_LIMIT:
                r2_last = MAX_ROW_LIMIT
            
            if r1_last < r1_first or r2_last < r2_first:
                return [], "Дані на аркушах виглядають порожніми"
            
            print(f"📊 Діапазони:")
            print(f"   {S1_NAME}: рядки {r1_first}-{r1_last}")
            print(f"   {S2_NAME}: рядки {r2_first}-{r2_last} (ліміт: {MAX_ROW_LIMIT})")
            
            # Будуємо індекси
            # ЗС: фільтр по F="БЗ", ігнор токенів у D (ПІБ)
            idx1 = self._build_index(
                self.ws1, r1_first, r1_last,
                key_col=S1_COL_KEY,
                subunit_col=S1_COL_KEY,  # D - ПІБ, застосовуємо ігнор токенів до ПІБ
                f_col=S1_COL_F, f_required=F_REQUIRED,
                g_col=None, g_ignore=None
            )
            
            # БЗ: без фільтра F, ігнор токенів у E (ПІБ), ігнор H містить "придані"
            idx2 = self._build_index(
                self.ws2, r2_first, r2_last,
                key_col=S2_COL_KEY,
                subunit_col=S2_COL_KEY,  # E - ПІБ, застосовуємо ігнор токенів до ПІБ
                f_col=None, f_required=None,
                g_col=S2_COL_H, g_ignore=H_IGNORE_TOKEN
            )
            
            print(f"📊 Індекси побудовано:")
            print(f"   {S1_NAME}: {len(idx1)} ключів")
            print(f"   {S2_NAME}: {len(idx2)} ключів")
            
            # Порівнюємо
            mismatches = []
            
            # ЗС -> БЗ (є в ЗС з міткою F='БЗ', але немає в БЗ)
            for key in idx1.keys():
                if key not in idx2:
                    for entry in idx1[key]:
                        entry.reason = 'в ЗС мітка F="БЗ", але відсутнє в листі БЗ'
                        mismatches.append(entry)
            
            # БЗ -> ЗС (є в БЗ, але немає в ЗС з міткою F='БЗ')
            for key in idx2.keys():
                if key not in idx1:
                    for entry in idx2[key]:
                        entry.reason = 'є в БЗ, але в ЗС відсутня мітка F="БЗ"'
                        mismatches.append(entry)
            
            return mismatches, ""
            
        except Exception as e:
            return [], f"Помилка: {str(e)}"
    
    def _find_sheet(self, name: str) -> Optional[Worksheet]:
        """Знайти лист за назвою (з урахуванням пробілів)"""
        name_norm = name.strip().lower()
        for ws in self.wb.worksheets:
            if ws.title.strip().lower() == name_norm:
                return ws
        return None
    
    def _max_last_row(self, ws: Worksheet, cols: List[int]) -> int:
        """Максимальний останній рядок серед кількох колонок"""
        max_row = 0
        for col in cols:
            # Шукаємо останню заповнену клітинку
            for row in range(ws.max_row, 0, -1):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip() != "":
                    if row > max_row:
                        max_row = row
                    break
        return max_row if max_row > 0 else 4
    
    def _build_index(
        self,
        ws: Worksheet,
        first_row: int,
        last_row: int,
        key_col: int,
        subunit_col: Optional[int] = None,
        f_col: Optional[int] = None,
        f_required: Optional[str] = None,
        g_col: Optional[int] = None,
        g_ignore: Optional[str] = None
    ) -> Dict[str, List[MismatchEntry]]:
        """
        Будує словник key -> List[MismatchEntry]
        
        Args:
            ws: Worksheet
            first_row, last_row: Діапазон рядків
            key_col: Колонка ключа (ПІБ)
            subunit_col: Колонка для перевірки IGNORE_TOKENS (може бути ПІБ або підрозділ)
            f_col, f_required: Фільтр по колонці F (якщо задано)
            g_col, g_ignore: Ігнор по колонці (H для БЗ) якщо містить певний текст
        """
        index = {}
        f_required_norm = self._normalize(f_required) if f_required else None
        g_ignore_norm = self._normalize(g_ignore) if g_ignore else None
        
        for row in range(first_row, last_row + 1):
            key_cell = ws.cell(row, key_col)
            key = self._normalize(key_cell.value)
            
            if not key:
                continue
            
            # Ігноруємо токени зі списку IGNORE_TOKENS
            # Для ЗС: перевіряємо D (ПІБ)
            # Для БЗ: перевіряємо E (ПІБ)
            if subunit_col is not None:
                check_value = self._normalize(ws.cell(row, subunit_col).value)
                if self._has_ignored_token(check_value):
                    continue
            
            # F-фільтр (якщо задано)
            passes_f = True
            if f_col is not None and f_required_norm:
                f_val = self._normalize(ws.cell(row, f_col).value)
                passes_f = (f_val == f_required_norm)
            
            # G-ігнор (якщо задано)
            passes_g = True
            if g_col is not None and g_ignore_norm:
                g_val = self._normalize(ws.cell(row, g_col).value)
                # Виключаємо рядок, якщо G містить цей підрядок
                passes_g = (g_ignore_norm not in g_val)
            
            if passes_f and passes_g:
                cell_addr = key_cell.coordinate
                value = str(key_cell.value) if key_cell.value else ""
                
                # Визначаємо причину невідповідності (буде додано пізніше)
                if key not in index:
                    index[key] = []
                
                # Створюємо запис (reason буде встановлено при порівнянні)
                entry = MismatchEntry(ws.title, cell_addr, value, "")
                index[key].append(entry)
        
        return index
    
    def _normalize(self, value) -> str:
        """Нормалізація значення (як у VBA Norm)"""
        if value is None or value == "":
            return ""
        
        s = str(value)
        
        # NBSP -> space
        s = s.replace('\u00A0', ' ')
        s = s.strip()
        
        # Множинні пробіли -> один
        while '  ' in s:
            s = s.replace('  ', ' ')
        
        return s.upper()
    
    def _has_ignored_token(self, normalized_text: str) -> bool:
        """
        Перевірка чи текст ТОЧНО співпадає з токеном для ігнорування
        Перевіряє ТОЧНУ відповідність, а не підрядок
        """
        text_normalized = normalized_text.replace(' ', '').upper()
        
        for token in IGNORE_TOKENS:
            token_normalized = token.upper().replace(' ', '')
            # ТОЧНА відповідність
            if text_normalized == token_normalized:
                return True
        
        return False


def check_mismatches(workbook_path: str) -> Tuple[List[MismatchEntry], str, Dict]:
    """
    Перевірити невідповідності між листами ЗС та БЗ
    
    Args:
        workbook_path: Шлях до Excel файлу
    
    Returns:
        (список_невідповідностей, повідомлення_про_помилку, статистика)
    """
    print(f"\n🔍 === ПЕРЕВІРКА НЕВІДПОВІДНОСТЕЙ ===")
    print(f"   Файл: {workbook_path}")
    
    checker = ExcelMismatchChecker(workbook_path)
    mismatches, error = checker.check()
    
    if error:
        print(f"❌ Помилка: {error}")
        return [], error, {}
    
    # Підраховуємо статистику (причини вже встановлені в check())
    s1_mismatches = [m for m in mismatches if m.sheet == S1_NAME]
    s2_mismatches = [m for m in mismatches if m.sheet == S2_NAME]
    
    stats = {
        'total': len(mismatches),
        's1_missing_in_s2': len(s1_mismatches),
        's2_missing_in_s1': len(s2_mismatches)
    }
    
    print(f"\n📊 Результати:")
    print(f"   Всього невідповідностей: {stats['total']}")
    print(f"   {S1_NAME} -> відсутні в {S2_NAME}: {stats['s1_missing_in_s2']}")
    print(f"   {S2_NAME} -> відсутні в {S1_NAME}: {stats['s2_missing_in_s1']}")
    
    if mismatches:
        print(f"\n📋 Превʼю (перші {min(PREVIEW_MAX, len(mismatches))}):")
        for i, m in enumerate(mismatches[:PREVIEW_MAX]):
            print(f"   {i+1}. {m}")
        
        if len(mismatches) > PREVIEW_MAX:
            print(f"   ... (ще {len(mismatches) - PREVIEW_MAX} записів)")
    
    return mismatches, "", stats
