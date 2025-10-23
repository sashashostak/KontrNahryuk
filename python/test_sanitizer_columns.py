"""
Тест санітизації тільки колонок A:H
"""

from openpyxl import Workbook
from excel_sanitizer import sanitize_cells

def test_column_range():
    """Тестування санітизації тільки колонок A:H"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Тест"
    
    # Додаємо дані в різні колонки
    # Колонка A (1)
    ws['A1'] = '25.08.2025'
    ws['A2'] = '  test  '
    
    # Колонка H (8)
    ws['H1'] = '01.12.2024'
    ws['H2'] = '1 234,56'
    
    # Колонка I (9) - поза межами A:H
    ws['I1'] = '15.03.2025'
    ws['I2'] = '  should not change  '
    
    # Колонка J (10) - поза межами A:H
    ws['J1'] = '25082025'
    
    print("🧪 === ТЕСТ САНІТИЗАЦІЇ КОЛОНОК A:H ===\n")
    print("📊 Дані ДО санітизації:")
    print(f"   A1: '{ws['A1'].value}' (DMY дата)")
    print(f"   A2: '{ws['A2'].value}' (текст з пробілами)")
    print(f"   H1: '{ws['H1'].value}' (DMY дата)")
    print(f"   H2: '{ws['H2'].value}' (число)")
    print(f"   I1: '{ws['I1'].value}' (дата, НЕ має змінитись)")
    print(f"   I2: '{ws['I2'].value}' (текст, НЕ має змінитись)")
    print(f"   J1: '{ws['J1'].value}' (дата, НЕ має змінитись)")
    
    # Збираємо тільки клітинки A:H
    cells = []
    for row in ws.iter_rows(min_col=1, max_col=8):  # A=1, H=8
        for cell in row:
            cells.append(cell)
    
    # Санітизуємо
    total, changed, preview = sanitize_cells(cells, show_preview=True)
    
    print(f"\n📈 Результати:")
    print(f"   Всього оброблено: {total} клітинок")
    print(f"   Змінено: {changed} клітинок")
    
    print(f"\n📊 Дані ПІСЛЯ санітизації:")
    print(f"   A1: '{ws['A1'].value}' (тип: {type(ws['A1'].value).__name__})")
    print(f"   A2: '{ws['A2'].value}' (тип: {type(ws['A2'].value).__name__})")
    print(f"   H1: '{ws['H1'].value}' (тип: {type(ws['H1'].value).__name__})")
    print(f"   H2: '{ws['H2'].value}' (тип: {type(ws['H2'].value).__name__})")
    print(f"   I1: '{ws['I1'].value}' (тип: {type(ws['I1'].value).__name__}) ← НЕ змінено")
    print(f"   I2: '{ws['I2'].value}' (тип: {type(ws['I2'].value).__name__}) ← НЕ змінено")
    print(f"   J1: '{ws['J1'].value}' (тип: {type(ws['J1'].value).__name__}) ← НЕ змінено")
    
    # Перевірка
    print(f"\n✅ Перевірка:")
    if ws['I1'].value == '15.03.2025' and ws['I2'].value == '  should not change  ' and ws['J1'].value == '25082025':
        print("   ✅ Колонки I та J НЕ змінені (правильно!)")
    else:
        print("   ❌ ПОМИЛКА: Колонки I та J були змінені!")
    
    if ws['A2'].value == 'test':
        print("   ✅ Колонка A очищена від пробілів (правильно!)")
    else:
        print(f"   ❌ ПОМИЛКА: A2 = '{ws['A2'].value}'")

if __name__ == '__main__':
    test_column_range()
