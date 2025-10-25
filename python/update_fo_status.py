#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
update_fo_status.py - Оновлення статусу FO (Формування особового складу)

Адаптація VBA логіки UpdateFO_Status_DOD10 для Python.
Записує звіт у аркуш LOG замість створення окремого аркуша.

Функціонал:
1. Читання значень з трьох зон (J:BI, BJ:ES, EU:FN)
2. Визначення статусу на основі "1" у колонках
3. Оновлення колонки FO
4. Запис детального звіту в LOG
"""

import re
from typing import Dict, List, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============== КОНСТАНТИ ==============

ROW_START = 10
ROW_END_MAX = 900

# Зони (номери колонок Excel, 1-based)
Z1_LEFT = 10   # J
Z1_RIGHT = 61  # BI

Z2_LEFT = 62   # BJ
Z2_RIGHT = 149 # ES

Z3_LEFT = 151  # EU
Z3_RIGHT = 170 # FN

# Префікси для зон
ZONE1_PREFIX = "Позиція"
ZONE2_PREFIX = "Краматорський р-н"


# ============== ДОПОМІЖНІ ФУНКЦІЇ ==============

def is_one(value) -> bool:
    """Перевіряє чи значення дорівнює 1"""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value == 1.0
    return str(value).strip() == "1"


def nz_str(value) -> str:
    """Повертає порожній рядок для None, інакше string"""
    if value is None:
        return ""
    return str(value)


def trim_spaces(s: str) -> str:
    """Нормалізує пробіли"""
    s = s.replace('\xa0', ' ')
    s = s.replace('\u202F', ' ')
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def header_at(ws: Worksheet, col: int, row: int = 5) -> str:
    """
    Повертає заголовок колонки з рядка 5 (як у VBA).
    """
    cell = ws.cell(row, col)
    
    # Перевірка об'єднаних комірок
    if cell.coordinate in ws.merged_cells:
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                top_left = ws.cell(merged.min_row, merged.min_col)
                return trim_spaces(nz_str(top_left.value))
    
    return trim_spaces(nz_str(cell.value))


def build_zone3_map(ws: Worksheet) -> Dict[int, str]:
    """
    Будує мапінг колонок Зони 3 -> готове значення FO.
    Точна копія VBA BuildZone3ColValueMap.
    """
    from openpyxl.utils import column_index_from_string
    
    mapping = {}
    
    # Точні адреси з VBA
    mapping[column_index_from_string('EV')] = "Лікування. До 10 діб"          # EV4
    mapping[column_index_from_string('EW')] = "Лікування. Понад 10 діб"      # EW4
    mapping[column_index_from_string('EU')] = "ВЛК"                           # EU3
    mapping[column_index_from_string('EY')] = "Відпустка щорічна"             # EY3
    mapping[column_index_from_string('EX')] = "Відпустка за станом здоров'я"  # EX3
    mapping[column_index_from_string('EZ')] = "НЦПП"                          # EZ4
    mapping[column_index_from_string('FN')] = "ППД. Інше"                     # FN4
    mapping[column_index_from_string('FF')] = "СЗЧ"                           # FF3
    mapping[column_index_from_string('FE')] = "Безвісті зниклі/загиблі"       # FE3
    mapping[column_index_from_string('FC')] = "Відрядження. Інше"             # FC4
    mapping[column_index_from_string('FD')] = "Арешт"                         # FD3
    mapping[column_index_from_string('FI')] = "ППД. Охорона складів, майна"   # FI4
    mapping[column_index_from_string('FA')] = "Відрядження. За кордоном"      # FA4
    mapping[column_index_from_string('FL')] = "ППД. Підготовка о/с"           # FL4
    mapping[column_index_from_string('FJ')] = "ППД. Господарські роботи"     # FJ4
    mapping[column_index_from_string('FB')] = "Відрядження. Інші в/ч"         # FB4
    mapping[column_index_from_string('FG')] = "Відмовники"                    # FG3
    mapping[column_index_from_string('FH')] = "ППД. Чергування, наряд"        # FH4
    mapping[column_index_from_string('FK')] = "ППД. Управління"               # FK4
    mapping[column_index_from_string('FM')] = "ППД. Адміністративна робота"   # FM4
    
    return mapping


# ============== ОСНОВНА ЛОГІКА ==============

def update_fo_status(ws: Worksheet, log_ws: Optional[Worksheet] = None, log_start_row: int = 2) -> Dict:
    """
    Оновлює колонку FO на основі значень у трьох зонах.
    Записує звіт у LOG worksheet.
    
    Returns:
        dict: {updates: int, errors: int, skipped: int, log_end_row: int}
    """
    
    # Знайдемо останній зайнятий рядок
    col_e = 5
    row_end = ROW_START
    for r in range(ROW_END_MAX, ROW_START - 1, -1):
        if ws.cell(r, col_e).value:
            row_end = r
            break
    
    if row_end < ROW_START:
        print(f"⚠️ У колонці E немає даних нижче рядка {ROW_START}")
        return {"updates": 0, "errors": 0, "skipped": 0, "log_end_row": log_start_row}
    
    print(f"📊 Обробка рядків {ROW_START}..{row_end} для оновлення статусу FO")
    
    # Знаходимо колонку FO (це колонка з літерами "FO" в Excel, не заголовок)
    from openpyxl.utils import column_index_from_string
    try:
        fo_col = column_index_from_string('FO')  # FO = 171
    except:
        fo_col = None
    
    if not fo_col:
        print("❌ Колонка FO не знайдена в рядку заголовків")
        
        # Записуємо повідомлення в LOG
        if log_ws:
            log_ws.cell(log_start_row, 1, '🔄 ОНОВЛЕННЯ СТАТУСУ FO')
            log_ws.cell(log_start_row, 1).font = Font(bold=True, size=12, color='FFFFFF')
            log_ws.cell(log_start_row, 1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            log_ws.merge_cells(f'A{log_start_row}:B{log_start_row}')
            log_start_row += 1
            
            log_ws.cell(log_start_row, 1, '❌ Колонка FO не знайдена в таблиці. Оновлення неможливе.')
            log_ws.cell(log_start_row, 1).font = Font(bold=True, color='C00000')
            log_ws.merge_cells(f'A{log_start_row}:B{log_start_row}')
            log_start_row += 1
            
            log_ws.cell(log_start_row, 1, 'Перевірте, що в цільовому файлі є колонка з назвою "FO" або "ФО".')
            log_ws.cell(log_start_row, 1).font = Font(italic=True, color='7F7F7F')
            log_ws.merge_cells(f'A{log_start_row}:B{log_start_row}')
            log_start_row += 1
        
        return {"updates": 0, "errors": 0, "skipped": 0, "log_end_row": log_start_row}
    
    print(f"✅ Колонка FO: {get_column_letter(fo_col)} (#{fo_col})")
    
    # Мапінг Зони 3
    z3_map = build_zone3_map(ws)
    print(f"📋 Мапінг Зони 3: {len(z3_map)} колонок")
    
    # Лічильники та зміни
    stats = {"updates": 0, "errors": 0, "skipped": 0}
    changes_log = []
    
    # Обробка рядків
    for row in range(ROW_START, row_end + 1):
        dept = nz_str(ws.cell(row, 2).value)
        pib = nz_str(ws.cell(row, 5).value)
        old_fo = nz_str(ws.cell(row, fo_col).value)
        
        # Пропускаємо порожні рядки (немає ПІБ)
        if not pib or not pib.strip():
            continue
        
        # Збір "1" з усіх зон
        zones = []
        cols = []
        
        for col in range(Z1_LEFT, Z1_RIGHT + 1):
            if is_one(ws.cell(row, col).value):
                zones.append(1)
                cols.append(col)
        
        for col in range(Z2_LEFT, Z2_RIGHT + 1):
            if is_one(ws.cell(row, col).value):
                zones.append(2)
                cols.append(col)
        
        for col in range(Z3_LEFT, Z3_RIGHT + 1):
            if is_one(ws.cell(row, col).value):
                zones.append(3)
                cols.append(col)
        
        found_count = len(zones)
        new_fo = ""
        typ = ""
        src_joined = "—"
        
        if found_count == 1:
            one_col = cols[0]
            one_header = header_at(ws, one_col)
            
            if zones[0] == 1:
                new_fo = f"{ZONE1_PREFIX}, {one_header}"
                typ = "Оновлено"
                src_joined = one_header
            elif zones[0] == 2:
                new_fo = f"{ZONE2_PREFIX}, {one_header}"
                typ = "Оновлено"
                src_joined = one_header
            elif zones[0] == 3:
                if one_col in z3_map:
                    new_fo = z3_map[one_col]
                    typ = "Оновлено"
                    src_joined = one_header
                else:
                    typ = "Помилка: немає правила для колонки Зона3"
                    src_joined = one_header
                    new_fo = ""  # Не пишемо нічого, якщо немає правила
        
        elif found_count == 0:
            typ = "Помилка: немає '1'"
        
        else:
            typ = "Помилка: кілька '1'"
            src_joined = "; ".join(header_at(ws, c) for c in cols)
        
        # Оновлення FO
        ws.cell(row, fo_col).value = new_fo if new_fo else None
        
        # Зберігаємо зміни
        if typ:
            if (typ == "Оновлено" and new_fo != old_fo) or (typ != "Оновлено"):
                changes_log.append({
                    'row': row,
                    'dept': dept,
                    'pib': pib,
                    'old_fo': old_fo,
                    'new_fo': new_fo,
                    'typ': typ,
                    'src': src_joined
                })
                
                if typ == "Оновлено":
                    stats["updates"] += 1
                else:
                    stats["errors"] += 1
        else:
            stats["skipped"] += 1
    
    # Запис у LOG (завжди, навіть якщо немає змін)
    log_row = log_start_row
    if log_ws:
        # Заголовок секції
        log_ws.cell(log_row, 1, '🔄 ОНОВЛЕННЯ СТАТУСУ FO')
        log_ws.cell(log_row, 1).font = Font(bold=True, size=12, color='FFFFFF')
        log_ws.cell(log_row, 1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        log_ws.cell(log_row, 1).alignment = Alignment(horizontal='left', vertical='center')
        log_ws.row_dimensions[log_row].height = 25
        log_ws.merge_cells(f'A{log_row}:B{log_row}')
        log_row += 1
        
        # Підсумок
        summary = f'Оновлено: {stats["updates"]}, Помилок: {stats["errors"]}, Пропущено: {stats["skipped"]}'
        log_ws.cell(log_row, 1, summary)
        log_ws.cell(log_row, 1).font = Font(bold=True, color='228B22' if stats["errors"] == 0 else '7F7F7F')
        log_ws.merge_cells(f'A{log_row}:B{log_row}')
        log_row += 1
        
        log_row += 1  # Порожній рядок
        
        # Таблиця змін (якщо є)
        if changes_log:
            # Заголовок таблиці
            log_ws.cell(log_row, 1, 'Рядок | Під. | ПІБ')
            log_ws.cell(log_row, 2, 'Було → Стало | Тип | Джерело')
            for col in [1, 2]:
                log_ws.cell(log_row, col).font = Font(bold=True, size=10)
                log_ws.cell(log_row, col).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            log_row += 1
            
            # Записуємо зміни (максимум 50)
            for change in changes_log[:50]:
                col_a = f"{change['row']} | {change['dept']} | {change['pib']}"
                col_b = f"{change['old_fo']} → {change['new_fo']} | {change['typ']} | {change['src']}"
                
                log_ws.cell(log_row, 1, col_a)
                log_ws.cell(log_row, 2, col_b)
                
                if change['typ'] == 'Оновлено':
                    fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    font = Font(color='228B22')
                else:
                    fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    font = Font(color='C00000')
                
                for col in [1, 2]:
                    log_ws.cell(log_row, col).fill = fill
                    log_ws.cell(log_row, col).font = font
                    log_ws.cell(log_row, col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                log_row += 1
            
            if len(changes_log) > 50:
                log_ws.cell(log_row, 1, f'... та ще {len(changes_log) - 50} змін')
                log_ws.cell(log_row, 1).font = Font(italic=True, color='7F7F7F')
                log_ws.merge_cells(f'A{log_row}:B{log_row}')
                log_row += 1
        else:
            # Якщо змін немає
            log_ws.cell(log_row, 1, 'Змін не виявлено. Всі значення FO актуальні.')
            log_ws.cell(log_row, 1).font = Font(italic=True, color='7F7F7F')
            log_ws.merge_cells(f'A{log_row}:B{log_row}')
            log_row += 1
    
    stats['log_end_row'] = log_row
    
    print(f"✅ Оновлення FO завершено:")
    print(f"   Оновлено: {stats['updates']}")
    print(f"   Помилок: {stats['errors']}")
    print(f"   Пропущено: {stats['skipped']}")
    
    return stats
