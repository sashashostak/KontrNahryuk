"""
Тест форматування листа LOG з розділювачами
"""

from openpyxl import Workbook
from excel_processor import ProcessingLog
from openpyxl.styles import Font, PatternFill

def test_log_formatting():
    """Тестування форматування LOG з розділювачами"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "LOG_TEST"
    
    # Створюємо тестовий лог
    log = ProcessingLog()
    
    # Додаємо записи
    log.add_separator("СПЕЦІАЛЬНА ОБРОБКА: 3БСП")
    log.add_entry("Копіювання діапазону (3БСП)", "файл1.xlsx", "3бСпП БЗ", 1368, "C4:H231")
    
    log.add_separator("ОБРОБКА ЛИСТА: ЗС")
    log.add_entry("Копіювання з файлу", "файл2.xlsx", "ЗС", 220, "Підрозділ: 1РСпП")
    log.add_entry("Копіювання з файлу", "файл3.xlsx", "ЗС", 220, "Підрозділ: 2РСпП")
    
    log.add_separator("ОБРОБКА ЛИСТА: БЗ")
    log.add_entry("Копіювання з файлу", "файл2.xlsx", "БЗ", 200, "Підрозділ: 1РСпП")
    log.add_entry("Копіювання з файлу", "файл3.xlsx", "БЗ", 33, "Підрозділ: 2РСпП")
    
    log.add_separator("САНІТИЗАЦІЯ ДАНИХ")
    log.add_entry("Санітизація даних (A:H)", "", "ЗС", 34, "Очищено 34 клітинок з 9440")
    log.add_entry("Санітизація даних (A:H)", "", "БЗ", 18, "Очищено 18 клітинок з 8144")
    
    log.add_separator("ПЕРЕВІРКА НЕВІДПОВІДНОСТЕЙ")
    log.add_entry("Перевірка невідповідностей", "", "ЗС / БЗ", 5, "ЗС→БЗ: 3, БЗ→ЗС: 2")
    
    print("🧪 === ТЕСТ ФОРМАТУВАННЯ LOG ===\n")
    print(f"📊 Всього записів: {len(log.entries)}")
    
    # Рахуємо розділювачі
    separators = sum(1 for e in log.entries if e.get('is_separator', False))
    regular = len(log.entries) - separators
    
    print(f"   Розділювачів: {separators}")
    print(f"   Звичайних записів: {regular}")
    
    print(f"\n📋 Превʼю записів:")
    entry_number = 0
    for entry in log.entries:
        if entry.get('is_separator', False):
            print(f"   [РОЗДІЛЮВАЧ] {entry['operation']}")
        else:
            entry_number += 1
            print(f"   {entry_number}. {entry['operation']} - {entry['sheet']} ({entry['rows']} рядків)")
    
    # Створюємо заголовки
    row = 1
    headers = ['№', 'Операція', 'Файл джерело', 'Лист', 'Рядків', 'Деталі']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx).value = header
    
    # Заповнюємо дані
    entry_number = 0
    for entry in log.entries:
        row += 1
        
        if entry.get('is_separator', False):
            # Розділювач
            cell = ws.cell(row=row, column=2)
            cell.value = entry['operation']
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws.merge_cells(f'B{row}:F{row}')
            ws.cell(row=row, column=1).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        else:
            # Звичайний запис
            entry_number += 1
            ws.cell(row=row, column=1).value = entry_number
            ws.cell(row=row, column=2).value = entry['operation']
            ws.cell(row=row, column=3).value = entry['source_file']
            ws.cell(row=row, column=4).value = entry['sheet']
            ws.cell(row=row, column=5).value = entry['rows']
            ws.cell(row=row, column=6).value = entry['details']
    
    # Зберігаємо тестовий файл
    test_file = "d:/Додатки/1SForS1/python/test_log_format.xlsx"
    wb.save(test_file)
    
    print(f"\n✅ Тестовий файл збережено: {test_file}")
    print(f"📂 Відкрийте файл для перегляду форматування")

if __name__ == '__main__':
    test_log_formatting()
