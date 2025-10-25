"""
Excel Slice Check - Перевірка колонок F та G
Адаптація VBA логіки для перевірки "зрізів"

Правила:
- Якщо в колонці F є токен (Ш, ВЛК, ВД), то колонка G має бути заповнена
- Перевіряються тільки рядки з ПІБ (колонка D)
- Структура ЗС: B=підрозділ, C=звання, D=ПІБ, E=псевдо, F=статус, G=дод.
"""

from typing import List, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from text_utils import normalize_text

# ========= НАЛАШТУВАННЯ =========
SOURCE_SHEET_NAME = "ЗС"    # Аркуш для перевірки

# Структура ЗС: B=підрозділ, C=звання, D=ПІБ, E=псевдо, F=статус, G=дод.
COL_SUBUNIT = 2             # B - підрозділ
COL_RANK = 3                # C - звання
COL_FIO = 4                 # D - ПІБ
COL_PSEUDO = 5              # E - псевдо
COL_F = 6                   # F - статус (де шукаємо токени Ш, ВЛК, ВД)
COL_G = 7                   # G - дод. (має бути заповнена якщо є токен)

# Токени у колонці F, що вимагають заповнення G
CHECK_TOKENS = ["Ш", "ВЛК", "ВД"]

PREVIEW_MAX = 25


class SliceIssue:
    """Запис про проблему з "зрізом" """
    
    def __init__(self, subunit: str, row: int, fio: str, pseudo: str, val_f: str, val_g: str, reason: str):
        self.subunit = subunit
        self.row = row
        self.fio = fio
        self.pseudo = pseudo
        self.val_f = val_f
        self.val_g = val_g
        self.reason = reason
    
    def __repr__(self):
        return f"SliceIssue(row={self.row}, fio={self.fio}, F={self.val_f})"


class SliceChecker:
    """Перевірка "зрізів" у колонках F/G"""
    
    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.wb = None
        self.ws = None
    
    def check(self) -> Tuple[List[SliceIssue], str]:
        """
        Перевірити аркуш на проблеми з "зрізами"
        
        Returns:
            (список_проблем, повідомлення_про_помилку)
        """
        try:
            self.wb = load_workbook(self.workbook_path)
            
            # Шукаємо аркуш
            self.ws = self._find_sheet(SOURCE_SHEET_NAME)
            if self.ws is None:
                return [], f"Не знайдено аркуш '{SOURCE_SHEET_NAME}'"
            
            # Визначаємо останній рядок
            last_row = self._get_last_row([COL_SUBUNIT, COL_FIO, COL_F, COL_G])
            if last_row < 1:
                return [], "Аркуш порожній"
            
            print(f"📊 Діапазон: рядки 1-{last_row}")
            
            # Перевіряємо кожен рядок
            issues = []
            for row in range(1, last_row + 1):
                fio_cell = self.ws.cell(row, COL_FIO)
                fio = self._normalize(fio_cell.value)
                
                # Пропускаємо рядки без ПІБ
                if not fio:
                    continue
                
                val_f = self._normalize(self.ws.cell(row, COL_F).value)
                val_g = self._normalize(self.ws.cell(row, COL_G).value)
                
                # Перевіряємо чи є токен у F
                if self._has_any_token(val_f) and not val_g:
                    subunit = self._normalize(self.ws.cell(row, COL_SUBUNIT).value)
                    pseudo = self._normalize(self.ws.cell(row, COL_PSEUDO).value)
                    
                    reason = f'У F="{val_f}", але G порожня — потрібен «зріз»'
                    
                    issue = SliceIssue(
                        subunit=subunit,
                        row=row,
                        fio=fio,
                        pseudo=pseudo,
                        val_f=val_f,
                        val_g="",
                        reason=reason
                    )
                    issues.append(issue)
            
            return issues, ""
            
        except Exception as e:
            return [], f"Помилка: {str(e)}"
    
    def _find_sheet(self, name: str) -> Optional[Worksheet]:
        """Знайти аркуш за назвою (з урахуванням пробілів)"""
        name_norm = name.strip().lower()
        for ws in self.wb.worksheets:
            if ws.title.strip().lower() == name_norm:
                return ws
        return None
    
    def _get_last_row(self, cols: List[int]) -> int:
        """Останній заповнений рядок серед кількох колонок"""
        max_row = 0
        for col in cols:
            for row in range(self.ws.max_row, 0, -1):
                cell = self.ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip() != "":
                    if row > max_row:
                        max_row = row
                    break
        return max_row if max_row > 0 else 0
    
    def _normalize(self, value) -> str:
        """Нормалізація значення - використовуємо єдину функцію"""
        # remove_spaces=False, щоб зберегти пробіли
        return normalize_text(value, remove_spaces=False)
    
    def _has_any_token(self, text: str) -> bool:
        """Перевірка чи текст містить будь-який токен"""
        if not text:
            return False

        # Нормалізуємо текст через єдину функцію
        text_normalized = normalize_text(text, remove_spaces=False)
        for token in CHECK_TOKENS:
            # Нормалізуємо токен через єдину функцію
            token_normalized = normalize_text(token, remove_spaces=False)
            if token_normalized in text_normalized:
                return True

        return False


def check_slices(workbook_path: str) -> Tuple[List[SliceIssue], str, dict]:
    """
    Перевірити "зрізи" (F/G) у файлі
    
    Args:
        workbook_path: Шлях до Excel файлу
    
    Returns:
        (список_проблем, повідомлення_про_помилку, статистика)
    """
    print(f"\n🔍 === ПЕРЕВІРКА «ЗРІЗІВ» (F/G) ===")
    print(f"   Файл: {workbook_path}")
    print(f"   Токени: {', '.join(CHECK_TOKENS)}")
    
    checker = SliceChecker(workbook_path)
    issues, error = checker.check()
    
    if error:
        print(f"❌ Помилка: {error}")
        return [], error, {}
    
    stats = {
        'total': len(issues)
    }
    
    print(f"\n📊 Результати:")
    print(f"   Всього проблем: {stats['total']}")
    
    if issues:
        print(f"\n📋 Превʼю (перші {min(len(issues), PREVIEW_MAX)}):")
        for i, issue in enumerate(issues[:PREVIEW_MAX], 1):
            print(f"   {i}. Рядок {issue.row} | {issue.subunit} | {issue.fio} | F={issue.val_f} | G=порожньо")
        
        if len(issues) > PREVIEW_MAX:
            print(f"   ... та ще {len(issues) - PREVIEW_MAX} проблем")
    
    return issues, "", stats


# Тестування
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        issues, error, stats = check_slices(sys.argv[1])
        if error:
            print(f"❌ {error}")
            sys.exit(1)
        else:
            print(f"\n✅ Перевірка завершена: {stats['total']} проблем")
    else:
        print("Використання: python excel_slice_check.py <шлях_до_файлу>")
