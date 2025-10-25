"""
Модуль для перевірки дублікатів в Excel файлах
Шукає задвоєння ПІБ в листах ЗС, БЗ та 3бСпП БЗ
"""

from typing import List, Dict, Tuple
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from text_utils import normalize_text


@dataclass
class DuplicateEntry:
    """Запис про дублікат"""
    sheet: str          # Назва листа
    column: str         # Колонка (D або E)
    value: str          # Значення, що дублюється
    rows: List[int]     # Список рядків, де знайдено дублікат
    count: int          # Кількість повторень


# Конфігурація листів та колонок для перевірки
SHEETS_CONFIG = {
    'ЗС': {
        'column': 'D',
        'col_idx': 4,
        'name': 'ЗС'
    },
    'БЗ': {
        'column': 'E',
        'col_idx': 5,
        'name': 'БЗ'
    },
    '3бСпП БЗ': {
        'column': 'E',
        'col_idx': 5,
        'name': '3бСпП БЗ'
    }
}


def _normalize_text(text: str) -> str:
    """Нормалізувати текст для порівняння - використовуємо єдину функцію"""
    # remove_spaces=False, щоб зберегти пробіли для ПІБ
    return normalize_text(text, remove_spaces=False)


def _find_duplicates_in_sheet(ws: Worksheet, col_idx: int, start_row: int = 2) -> Dict[str, List[int]]:
    """
    Знайти дублікати в колонці листа
    
    Args:
        ws: Лист Excel
        col_idx: Індекс колонки (1-based)
        start_row: З якого рядка починати (за замовчуванням 2 - пропускаємо заголовок)
    
    Returns:
        Словник {значення: [список_рядків]}
    """
    values_map = {}
    
    for row_idx in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        
        # Пропускаємо порожні клітинки
        if not value:
            continue
        
        # Нормалізуємо текст
        normalized = _normalize_text(value)
        
        if not normalized:
            continue
        
        # Додаємо до словника
        if normalized not in values_map:
            values_map[normalized] = []
        values_map[normalized].append(row_idx)
    
    # Повертаємо тільки дублікати (де більше 1 рядка)
    duplicates = {k: v for k, v in values_map.items() if len(v) > 1}
    
    return duplicates


def check_duplicates(workbook_path: str) -> Tuple[List[DuplicateEntry], str, Dict]:
    """
    Перевірити дублікати в файлі
    
    Args:
        workbook_path: Шлях до Excel файлу
    
    Returns:
        (список_дублікатів, повідомлення_про_помилку, статистика)
    """
    try:
        # Завантажуємо книгу
        wb = load_workbook(workbook_path, data_only=True)
        
        all_duplicates = []
        total_duplicates = 0
        stats_by_sheet = {}
        
        # Перевіряємо кожен лист
        for sheet_key, config in SHEETS_CONFIG.items():
            sheet_name = config['name']
            
            # Шукаємо лист
            if sheet_name not in wb.sheetnames:
                print(f"   ⚠️ Лист '{sheet_name}' не знайдено, пропускаємо")
                continue
            
            ws = wb[sheet_name]
            col_idx = config['col_idx']
            col_letter = config['column']
            
            print(f"   🔍 Перевірка листа '{sheet_name}', колонка {col_letter}...")
            
            # Шукаємо дублікати
            duplicates = _find_duplicates_in_sheet(ws, col_idx)
            
            sheet_dup_count = 0
            
            for value, rows in duplicates.items():
                # Отримуємо оригінальне значення (не нормалізоване) з першого рядка
                original_value = ws.cell(row=rows[0], column=col_idx).value
                
                entry = DuplicateEntry(
                    sheet=sheet_name,
                    column=col_letter,
                    value=str(original_value),
                    rows=rows,
                    count=len(rows)
                )
                all_duplicates.append(entry)
                sheet_dup_count += 1
            
            stats_by_sheet[sheet_name] = sheet_dup_count
            total_duplicates += sheet_dup_count
            
            if sheet_dup_count > 0:
                print(f"      ⚠️ Знайдено {sheet_dup_count} дублікатів")
            else:
                print(f"      ✅ Дублікатів не знайдено")
        
        wb.close()
        
        # Статистика
        stats = {
            'total': total_duplicates,
            'by_sheet': stats_by_sheet
        }
        
        return all_duplicates, "", stats
        
    except FileNotFoundError:
        return [], f"Файл не знайдено: {workbook_path}", {}
    except Exception as e:
        return [], f"Помилка при перевірці дублікатів: {str(e)}", {}


if __name__ == "__main__":
    # Тестування
    import sys
    
    if len(sys.argv) < 2:
        print("Використання: python excel_duplicates.py <шлях_до_файлу>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    duplicates, error, stats = check_duplicates(file_path)
    
    if error:
        print(f"❌ Помилка: {error}")
        sys.exit(1)
    
    print(f"\n📊 === РЕЗУЛЬТАТИ ПЕРЕВІРКИ ДУБЛІКАТІВ ===")
    print(f"Всього знайдено дублікатів: {stats['total']}")
    
    for sheet, count in stats['by_sheet'].items():
        print(f"  {sheet}: {count}")
    
    if duplicates:
        print(f"\n📝 Детальна інформація:")
        for dup in duplicates:
            rows_str = ", ".join(map(str, dup.rows))
            print(f"  [{dup.sheet}] Колонка {dup.column}: «{dup.value}» — рядки: {rows_str} ({dup.count} разів)")
